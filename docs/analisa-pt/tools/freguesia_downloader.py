#!/usr/bin/env python3
"""Freguesia Downloader — Download official Portuguese parish data with NIFs

Downloads the official Freguesiasdadosgerais.xlsx from dados.gov.pt which contains:
- NIF numbers for all Portuguese parishes (freguesias)
- Parish names, addresses, postal codes
- INE administrative codes
- Contact information

Usage:
    python freguesia_downloader.py download    # Download and parse the dataset
    python freguesia_downloader.py stats       # Show statistics
    python freguesia_downloader.py lookup 123  # Look up by NIF or name
"""

import sys
import json
import sqlite3
import argparse
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
OUTPUT_FILE = DATA_DIR / "freguesia_nif_database.json"
DB_PATH = DATA_DIR / "procurement.db"

# Official dados.gov.pt dataset with freguesia NIFs
DATASET_URL = "https://dados.gov.pt/s/dadosGovFiles/Freguesiasdadosgerais.xlsx"
LOCAL_FILE = DATA_DIR / "freguesiasdadosgerais.xlsx"

# Import CAOP codes for municipality resolution
try:
    from caop_codes import (
        CODE_TO_MUNICIPALITY, _normalize_name, resolve_municipality_6digit,
        DISTRICT_CODES, CODE_TO_DISTRICT
    )
except ImportError:
    CODE_TO_MUNICIPALITY = {}
    _normalize_name = lambda x: x.lower().strip()
    resolve_municipality_6digit = lambda x: None


def _check_openpyxl() -> bool:
    """Check if openpyxl is installed before downloading."""
    try:
        import openpyxl
        return True
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return False


def download_dataset() -> bool:
    """Download the freguesia dataset from dados.gov.pt."""
    if not _check_openpyxl():
        return False

    print(f"Downloading freguesia dataset from dados.gov.pt...")
    print(f"  URL: {DATASET_URL}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    try:
        urllib.request.urlretrieve(DATASET_URL, LOCAL_FILE)
        print(f"  Saved to: {LOCAL_FILE}")

        # Validate file is not an HTML error page
        file_size = LOCAL_FILE.stat().st_size
        if file_size < 1000:
            print(f"  Error: File too small ({file_size} bytes) — likely an error page")
            LOCAL_FILE.unlink()
            return False

        # Check XLSX signature (ZIP format starts with PK)
        with open(LOCAL_FILE, 'rb') as f:
            header = f.read(2)
        if header != b'PK':
            print(f"  Error: Not a valid XLSX file (got {header!r} instead of ZIP signature)")
            LOCAL_FILE.unlink()
            return False

        print(f"  Validated: {file_size:,} bytes")
        return True

    except Exception as e:
        print(f"  Error downloading: {e}")
        return False


def _find_column(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Find a column index by matching header names (case-insensitive)."""
    for i, header in enumerate(headers):
        h = str(header or "").strip().lower()
        for candidate in candidates:
            if candidate.lower() in h:
                return i
    return None


def parse_xlsx() -> Dict[str, Dict]:
    """Parse the XLSX file and extract freguesia data."""
    if not LOCAL_FILE.exists():
        print(f"Error: {LOCAL_FILE} not found. Run 'download' first.")
        return {}

    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return {}

    print(f"Parsing {LOCAL_FILE}...")
    wb = openpyxl.load_workbook(LOCAL_FILE, read_only=True)
    ws = wb.active

    # Get headers from first row
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  Found {len(headers)} columns")
    print(f"  Headers: {headers[:10]}")

    # Find column indices with multiple fallback candidates
    nif_col = _find_column(headers, ["nif", "número de identificação fiscal", "identificação fiscal"])
    name_col = _find_column(headers, ["freguesia", "nome", "designação", "descrição", "descrição da freguesia"])
    ine_col = _find_column(headers, ["ine", "código ine", "codigoine", "codigo_ine", "cod_ine"])
    muni_col = _find_column(headers, ["concelho", "município", "municipio", "código concelho", "codigo concelho"])
    dist_col = _find_column(headers, ["distrito", "código distrito", "codigo distrito"])

    if nif_col is None:
        print(f"  Error: Could not find NIF column. Headers: {headers}")
        wb.close()
        return {}

    print(f"  Using columns: NIF={nif_col}, Name={name_col}, INE={ine_col}, Municipality={muni_col}")

    # Parse rows
    freguesias = {}
    row_count = 0
    nif_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1

        nif = str(row[nif_col] or "").strip()
        if not nif or len(nif) != 9 or not nif.isdigit():
            continue

        nif_count += 1
        name = str(row[name_col] or "").strip() if name_col is not None else ""
        ine_code = str(row[ine_col] or "").strip() if ine_col is not None else ""
        municipality = str(row[muni_col] or "").strip() if muni_col is not None else ""
        district = str(row[dist_col] or "").strip() if dist_col is not None else ""

        # Resolve municipality from INE code if not provided directly
        if not municipality and ine_code and len(ine_code) >= 4:
            muni_code = ine_code[:4]
            if muni_code in CODE_TO_MUNICIPALITY:
                municipality = CODE_TO_MUNICIPALITY[muni_code]

        # Resolve district from INE code
        if not district and ine_code and len(ine_code) >= 2:
            dist_code = ine_code[:2]
            if dist_code in CODE_TO_DISTRICT:
                district = CODE_TO_DISTRICT[dist_code]

        freguesias[nif] = {
            "nif": nif,
            "name": name,
            "ine_code": ine_code,
            "municipality": municipality,
            "district": district,
        }

    wb.close()
    print(f"  Parsed {nif_count} parishes with valid NIFs from {row_count} rows")
    return freguesias


def save_database(freguesias: Dict[str, Dict]):
    """Save freguesia database to JSON."""
    output = {
        "version": "1.0",
        "source": "dados.gov.pt Freguesiasdadosgerais.xlsx",
        "total": len(freguesias),
        "freguesias": freguesias,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Saved {len(freguesias)} parishes to {OUTPUT_FILE}")


def cmd_download(args):
    """Download and parse freguesia dataset."""
    if not download_dataset():
        return

    freguesias = parse_xlsx()
    if not freguesias:
        print("Error: No valid parishes found in dataset")
        return

    save_database(freguesias)

    # Cross-validate with procurement.db
    _cross_validate_with_procurement(freguesias)

    # Auto-show stats
    cmd_stats(None)



def _cross_validate_with_procurement(freguesias: Dict[str, Dict]):
    """Check how many downloaded NIFs appear in the procurement database."""
    if not DB_PATH.exists():
        return

    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.execute("SELECT COUNT(DISTINCT adjudicante_nif) FROM contratos WHERE adjudicante_nif IN ({})".format(
            ','.join(['?'] * len(freguesias))))
        match_count = cursor.fetchone()[0]
        conn.close()

        if match_count > 0:
            print(f"  Cross-validation: {match_count} of {len(freguesias)} parishes have contracts in procurement.db")
    except Exception as e:
        print(f"  Cross-validation skipped: {e}")


def cmd_stats(args):
    """Show database statistics."""
    if not OUTPUT_FILE.exists():
        print("No database found. Run 'download' first.")
        return

    with open(OUTPUT_FILE) as f:
        data = json.load(f)

    freguesias = data.get("freguesias", {})

    # Count by municipality
    muni_counts = {}
    with_muni = 0
    with_ine = 0
    for nif, f in freguesias.items():
        muni = f.get("municipality", "")
        if muni:
            muni_counts[muni] = muni_counts.get(muni, 0) + 1
            with_muni += 1
        if f.get("ine_code"):
            with_ine += 1

    print(f"\n{'='*70}")
    print(f"  FREGUESIA NIF DATABASE — STATISTICS")
    print(f"{'='*70}")
    print(f"  Total parishes:           {len(freguesias):>10,}")
    print(f"  With municipality:        {with_muni:>10,} ({with_muni*100/len(freguesias):.1f}%)")
    print(f"  With INE code:            {with_ine:>10,} ({with_ine*100/len(freguesias):.1f}%)")
    print(f"  Municipalities covered:   {len(muni_counts):>10}")
    print(f"\n  Top 10 municipalities by parish count:")
    for muni, count in sorted(muni_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"    {muni:<35} {count:>5} parishes")
    print(f"{'='*70}\n")


def cmd_lookup(args):
    """Look up a parish by NIF or name."""
    if not OUTPUT_FILE.exists():
        print("No database found. Run 'download' first.")
        return

    with open(OUTPUT_FILE) as f:
        data = json.load(f)

    freguesias = data.get("freguesias", {})
    query = args.query.lower()

    # Search by NIF or name
    matches = []
    for nif, f in freguesias.items():
        if query in nif or query in f.get("name", "").lower():
            matches.append(f)

    if not matches:
        print(f"No parishes found matching '{args.query}'")
        return

    print(f"\n{'='*70}")
    print(f"  MATCHES ({len(matches)} parishes)")
    print(f"{'='*70}")

    for f in matches[:10]:
        print(f"\n  NIF: {f['nif']}")
        print(f"  Name: {f.get('name', '?')}")
        print(f"  Municipality: {f.get('municipality', '?')}")
        print(f"  District: {f.get('district', '?')}")
        print(f"  INE Code: {f.get('ine_code', '?')}")

    if len(matches) > 10:
        print(f"\n  ... and {len(matches) - 10} more")


def main():
    parser = argparse.ArgumentParser(
        description="Download official Portuguese freguesia data with NIFs",
    )
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("download", help="Download and parse the dataset")
    sub.add_parser("stats", help="Show database statistics")

    lookup_p = sub.add_parser("lookup", help="Look up a parish")
    lookup_p.add_argument("query", help="NIF or name to search")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "download": cmd_download,
        "stats": cmd_stats,
        "lookup": cmd_lookup,
    }
    commands[args.command](args)


if __name__ == "__main__":
    main()
