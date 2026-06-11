#!/usr/bin/env python3
"""Unified Procurement Database — Single SQLite for all procurement data.

Consolidates contratos (signed contracts) and entidades (entities) into
procurement.db, eliminating the need to re-parse XLSX files on every
tool invocation.

Usage:
    python procurement_db.py download            # Download XLSX from dados.gov.pt
    python procurement_db.py download --years 2024 2025  # Download specific years
    python procurement_db.py build               # Build/refresh the database
    python procurement_db.py build --force        # Clear and rebuild
    python procurement_db.py build --auto-download # Download if missing, then build
    python procurement_db.py stats                # Show database statistics
    python procurement_db.py query \"SELECT ...\"  # Run arbitrary SQL
"""

import sys
import json
import re
import sqlite3
import argparse
import subprocess
import time
from pathlib import Path
from datetime import datetime, timezone

try:
    import urllib.request
    import ssl
except ImportError:
    print("ERROR: urllib required (built-in)")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
DB_PATH = DATA_DIR / "procurement.db"
ENTIDADES_XLSX = DATA_DIR / "entidades.xlsx"
XLSX_DIR = DATA_DIR

# dados.gov.pt dataset IDs (CKAN API)
API_BASE = "https://dados.gov.pt/api/1/datasets/"
CONTRATOS_DATASET_ID = "66d72d488ca4b7cb2de28712"  # Contratos 2012-2026
ENTIDADES_DATASET_ID = "67d80b2c4750b888116940fb"  # Entidades

# SSL context
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/octet-stream,*/*",
}


# ---------------------------------------------------------------------------
# Database Schema
# ---------------------------------------------------------------------------

def init_db(force: bool = False) -> sqlite3.Connection:
    """Initialize the unified procurement database."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if force and DB_PATH.exists():
        DB_PATH.unlink()
        print("  Deleted existing database.")

    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-64000")  # 64MB cache

    # Contracts table — mirrors contratos2025.xlsx
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contratos (
            idcontrato INTEGER PRIMARY KEY,
            nAnuncio TEXT,
            tipoAnuncio TEXT,
            idINCM TEXT,
            tipoContrato TEXT,
            idprocedimento TEXT,
            tipoprocedimento TEXT,
            objectoContrato TEXT,
            descContrato TEXT,
            adjudicante TEXT,
            adjudicante_nif TEXT,
            adjudicante_nome TEXT,
            adjudicatarios TEXT,
            dataPublicacao TEXT,
            dataCelebracaoContrato TEXT,
            precoContratual REAL,
            CPV TEXT,
            prazoExecucao INTEGER,
            LocalExecucao TEXT,
            fundamentacao TEXT,
            ProcedimentoCentralizado TEXT,
            numAcordoQuadro TEXT,
            DescrAcordoQuadro TEXT,
            precoBaseProcedimento REAL,
            dataDecisaoAdjudicacao TEXT,
            dataFechoContrato TEXT,
            PrecoTotalEfetivo REAL,
            regime TEXT,
            justifNReducEscrContrato TEXT,
            tipoFimContrato TEXT,
            CritMateriais TEXT,
            concorrentes TEXT,
            linkPecasProc TEXT,
            Observacoes TEXT,
            ContratEcologico TEXT,
            fundamentAjusteDireto TEXT,
            Ano INTEGER,
            adjudicatarioPMEs TEXT,
            NUTs TEXT,
            Lotes TEXT,
            TipoCriterioAdjudicacao TEXT
        )
    """)

    # Entities table — mirrors entidades.xlsx
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entidades (
            nifEntidade TEXT PRIMARY KEY,
            desigEntidade TEXT,
            numContratos INTEGER,
            totAdjudicatario INTEGER,
            totValorContratIni REAL,
            totAdjudicante INTEGER,
            totAdjudicanteValorContratIni REAL,
            descPais TEXT,
            AliasPais TEXT
        )
    """)

    # Metadata table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT
        )
    """)

    # Indexes for common queries
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_nAnuncio ON contratos(nAnuncio)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_adjudicante_nif ON contratos(adjudicante_nif)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_tipo ON contratos(tipoContrato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_preco ON contratos(precoContratual)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_proc ON contratos(tipoprocedimento)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_data ON contratos(dataPublicacao)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_c_link ON contratos(linkPecasProc)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_e_pais ON entidades(descPais)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_e_num ON entidades(numContratos)")

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Download from dados.gov.pt
# ---------------------------------------------------------------------------

def api_get(dataset_id: str) -> dict:
    """Fetch dataset metadata from dados.gov.pt CKAN API."""
    url = f"{API_BASE}{dataset_id}"
    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=15, context=SSL_CTX)
    return json.loads(resp.read())


def get_resource_urls(dataset_id: str, fmt: str = "xlsx") -> dict[str, str]:
    """Get download URLs for a dataset's XLSX resources.

    Returns: {filename: url} mapping.
    """
    data = api_get(dataset_id)
    ds = data.get("data", data)
    urls = {}
    for r in ds.get("resources", []):
        r_fmt = str(r.get("format", "")).lower()
        if r_fmt == fmt.lower():
            name = r.get("title", r.get("name", ""))
            dl_url = r.get("url", "")
            if name and dl_url:
                urls[name] = dl_url
    return urls


def download_file(url: str, local_path: Path, label: str = "") -> bool:
    """Download a file from URL to local path. Returns True on success."""
    if local_path.exists() and local_path.stat().st_size > 1000:
        size = local_path.stat().st_size
        print(f"    {label}: already exists ({size:,} bytes)")
        return True

    print(f"    {label}: downloading...", end=" ", flush=True)
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=120, context=SSL_CTX)
        data = resp.read()
        local_path.write_bytes(data)
        print(f"done ({len(data):,} bytes)")
        return True
    except Exception as e:
        print(f"FAILED: {e}")
        return False


def cmd_download(args):
    """Download contratos and entidades XLSX files from dados.gov.pt."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Determine which years to download
    years = getattr(args, "years", None)
    entidades_only = getattr(args, "entidades_only", False)
    contratos_only = getattr(args, "contratos_only", False)

    total_files = 0
    total_size = 0

    # --- Download contratos ---
    if not entidades_only:
        print(f"\n  Downloading contratos from dados.gov.pt...")
        try:
            urls = get_resource_urls(CONTRATOS_DATASET_ID)
        except Exception as e:
            print(f"    ERROR fetching metadata: {e}")
            urls = {}

        if urls:
            # Filter by year if specified
            target_files = {}
            for name, url in urls.items():
                m = re.search(r"contratos(\d{4})", name)
                if m:
                    year = int(m.group(1))
                    if years and year not in years:
                        continue
                    target_files[name] = url

            print(f"    Found {len(target_files)} XLSX file(s)")
            for name, url in sorted(target_files.items()):
                local_path = XLSX_DIR / name
                if download_file(url, local_path, name):
                    total_files += 1
                    total_size += local_path.stat().st_size
                time.sleep(0.5)

    # --- Download entidades ---
    if not contratos_only:
        print(f"\n  Downloading entidades from dados.gov.pt...")
        try:
            urls = get_resource_urls(ENTIDADES_DATASET_ID)
        except Exception as e:
            print(f"    ERROR fetching metadata: {e}")
            urls = {}

        if urls:
            for name, url in urls.items():
                local_path = XLSX_DIR / name
                if download_file(url, local_path, name):
                    total_files += 1
                    total_size += local_path.stat().st_size
                time.sleep(0.5)

    print(f"\n  Total: {total_files} files, {total_size:,} bytes")
    if total_size > 0:
        print(f"  Run 'python procurement_db.py build' to build the database")
    print()


def has_any_contratos_xlsx() -> bool:
    """Check if any contratos*.xlsx file exists in the data directory."""
    return any(DATA_DIR.glob("contratos*.xlsx"))


def ensure_xlsx_files() -> bool:
    """Check if required XLSX files exist. Returns True if ready to build."""
    missing = []
    if not has_any_contratos_xlsx():
        missing.append("contratos*.xlsx")
    if not ENTIDADES_XLSX.exists():
        missing.append("entidades.xlsx")

    if not missing:
        return True

    print(f"  Missing XLSX files: {', '.join(missing)}")
    print(f"  Run 'python procurement_db.py download' first, or")
    print(f"  use 'build --auto-download' to download automatically")
    return False


# ---------------------------------------------------------------------------
# XLSX Parsers
# ---------------------------------------------------------------------------

def _parse_one_contratos_xlsx(conn: sqlite3.Connection, xlsx_path: Path, file_year: int) -> int:
    """Parse a single contratos XLSX file into the contratos table."""
    print(f"  {xlsx_path.name}: parsing...", end=" ", flush=True)
    t0 = time.time()

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    def get(row, field, default=""):
        idx = h.get(field, -1)
        if idx < 0 or idx >= len(row):
            return default
        val = row[idx]
        return val if val is not None else default

    def get_num(row, field):
        val = get(row, field, None)
        if val is None:
            return None
        try:
            v = float(val)
            if v > 1e10:
                return None
            return v
        except (ValueError, TypeError):
            return None

    def fmt_date(val):
        if val is None:
            return ""
        if hasattr(val, 'strftime'):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()[:10]

    def extract_nif(adjudicante):
        """Extract NIF from adjudicante field (format: 'NIF - Entity Name')."""
        s = str(adjudicante or "")
        if " - " in s:
            parts = s.split(" - ", 1)
            return parts[0].strip(), parts[1].strip()
        return "", s

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        idcontrato = get(row, "idcontrato")
        try:
            idcontrato = int(idcontrato)
        except (ValueError, TypeError):
            continue

        adjudicante = str(get(row, "adjudicante"))
        nif, nome = extract_nif(adjudicante)

        rows.append((
            idcontrato,
            str(get(row, "nAnuncio")).strip(),
            str(get(row, "tipoAnuncio")).strip(),
            str(get(row, "idINCM")).strip(),
            str(get(row, "tipoContrato")).strip(),
            str(get(row, "idprocedimento")).strip(),
            str(get(row, "tipoprocedimento")).strip(),
            str(get(row, "objectoContrato")).strip()[:2000],
            str(get(row, "descContrato")).strip()[:2000],
            adjudicante.strip()[:200],
            nif,
            nome.strip()[:200],
            str(get(row, "adjudicatarios")).strip()[:500],
            fmt_date(get(row, "dataPublicacao")),
            fmt_date(get(row, "dataCelebracaoContrato")),
            get_num(row, "precoContratual"),
            str(get(row, "CPV")).strip(),
            None,  # prazoExecucao -- not in all rows
            str(get(row, "LocalExecucao")).strip()[:200],
            str(get(row, "fundamentacao")).strip()[:500],
            str(get(row, "ProcedimentoCentralizado")).strip(),
            str(get(row, "numAcordoQuadro")).strip(),
            str(get(row, "DescrAcordoQuadro")).strip()[:200],
            get_num(row, "precoBaseProcedimento"),
            fmt_date(get(row, "dataDecisaoAdjudicacao")),
            fmt_date(get(row, "dataFechoContrato")),
            get_num(row, "PrecoTotalEfetivo"),
            str(get(row, "regime")).strip(),
            str(get(row, "justifNReducEscrContrato")).strip()[:500],
            str(get(row, "tipoFimContrato")).strip(),
            str(get(row, "CritMateriais")).strip(),
            str(get(row, "concorrentes")).strip(),
            str(get(row, "linkPecasProc")).strip()[:500],
            str(get(row, "Observacoes")).strip()[:500],
            str(get(row, "ContratEcologico")).strip(),
            str(get(row, "fundamentAjusteDireto")).strip()[:500],
            file_year,
            str(get(row, "adjudicatarioPMEs")).strip(),
            str(get(row, "NUTs")).strip(),
            str(get(row, "Lotes")).strip(),
            str(get(row, "TipoCriterioAdjudicacao")).strip(),
        ))

    wb.close()

    conn.executemany(
        "INSERT OR REPLACE INTO contratos VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    conn.commit()

    elapsed = time.time() - t0
    print(f"{len(rows):,} rows in {elapsed:.1f}s")
    return len(rows)


def parse_contratos_xlsx(conn: sqlite3.Connection) -> int:
    """Parse all contratos*.xlsx files into the contratos table."""
    xlsx_files = sorted(DATA_DIR.glob("contratos*.xlsx"))
    if not xlsx_files:
        print("  No contratos*.xlsx files found")
        return 0

    print(f"  Found {len(xlsx_files)} contratos XLSX file(s)")
    total = 0
    for xlsx_path in xlsx_files:
        m = re.search(r"contratos(\d{4})", xlsx_path.name)
        file_year = int(m.group(1)) if m else 0
        total += _parse_one_contratos_xlsx(conn, xlsx_path, file_year)

    # Add year index if missing
    try:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_c_ano ON contratos(Ano)")
        conn.commit()
    except Exception:
        pass

    return total


def parse_entidades_xlsx(conn: sqlite3.Connection) -> int:
    """Parse entidades.xlsx into the entidades table."""
    if not ENTIDADES_XLSX.exists():
        print(f"  ERROR: {ENTIDADES_XLSX} not found")
        return 0

    print(f"  Parsing entidades.xlsx...", end=" ", flush=True)
    t0 = time.time()

    wb = openpyxl.load_workbook(str(ENTIDADES_XLSX), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    def get(row, field, default=""):
        idx = h.get(field, -1)
        if idx < 0 or idx >= len(row):
            return default
        val = row[idx]
        return val if val is not None else default

    def get_num(row, field):
        val = get(row, field, None)
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        nif = str(get(row, "nifEntidade")).strip()
        if not nif or nif == "-":
            continue

        rows.append((
            nif,
            str(get(row, "desigEntidade")).strip()[:200],
            int(get_num(row, "numContratos") or 0),
            int(get_num(row, "totAdjudicatario") or 0),
            get_num(row, "totValorContratIni"),
            int(get_num(row, "totAdjudicante") or 0),
            get_num(row, "totAdjudicanteValorContratIni"),
            str(get(row, "descPais")).strip(),
            str(get(row, "AliasPais")).strip(),
        ))

    wb.close()

    conn.executemany(
        "INSERT OR REPLACE INTO entidades VALUES (?,?,?,?,?,?,?,?,?)",
        rows,
    )
    conn.commit()

    elapsed = time.time() - t0
    print(f"{len(rows):,} rows in {elapsed:.1f}s")
    return len(rows)


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_build(args):
    """Build/refresh the unified procurement database."""
    # Auto-download if requested and files are missing
    auto_dl = getattr(args, "auto_download", False)
    if auto_dl:
        missing = []
        if not has_any_contratos_xlsx():
            missing.append("contratos*.xlsx")
        if not ENTIDADES_XLSX.exists():
            missing.append("entidades.xlsx")
        if missing:
            print(f"  Auto-downloading missing files: {', '.join(missing)}")
            dl_args = argparse.Namespace(years=None, entidades_only=False, contratos_only=False)
            cmd_download(dl_args)

    if not ensure_xlsx_files():
        return

    print(f"\n  Building procurement.db...\n")
    conn = init_db(force=args.force)

    t0 = time.time()
    n_contratos = parse_contratos_xlsx(conn)
    n_entidades = parse_entidades_xlsx(conn)

    # Store metadata
    conn.execute(
        "INSERT OR REPLACE INTO meta (key, value, updated_at) VALUES (?, ?, ?)",
        ("last_build", f"{n_contratos} contratos, {n_entidades} entidades",
         datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    conn.close()

    elapsed = time.time() - t0
    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    print(f"\n  Done in {elapsed:.1f}s — {db_size:.1f} MB")
    print(f"  Database: {DB_PATH}\n")

    # Auto-build the procurement cache so the dashboard loads fast on next request
    cache_script = SCRIPT_DIR / "procurement_cache.py"
    if cache_script.exists():
        print(f"\n  Building procurement cache for fast dashboard loading...")
        try:
            result = subprocess.run(
                [sys.executable, str(cache_script), "build"],
                cwd=str(SCRIPT_DIR),
                check=False,
            )
            if result.returncode != 0:
                print(f"  ⚠️  Cache build exited with code {result.returncode}.")
                print(f"  Run manually: python procurement_cache.py build")
        except Exception as e:
            print(f"  Cache build skipped: {e}")
            print(f"  Run manually: python procurement_cache.py build")
    else:
        print(f"\n  (procurement_cache.py not found — skipping cache build)")


def cmd_status(args):
    """Quick one-glance status overview."""
    if not DB_PATH.exists():
        print(f"  procurement.db: NOT FOUND")
        print(f"  Run 'python procurement_db.py build' to create")
        return

    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    mtime = datetime.fromtimestamp(DB_PATH.stat().st_mtime)
    age_days = (datetime.now() - mtime).days

    conn = sqlite3.connect(str(DB_PATH))
    tables = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]

    counts = {}
    for t in tables:
        if t != "meta":
            counts[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]

    last_build = None
    try:
        row = conn.execute("SELECT value, updated_at FROM meta WHERE key='last_build'").fetchone()
        if row:
            last_build = row[1]
    except Exception:
        pass
    conn.close()

    print(f"  procurement.db  {db_size:.1f} MB  ({age_days}d old)")
    for t, c in counts.items():
        print(f"    {t:<20} {c:>10,}")
    if last_build:
        print(f"    last build: {last_build}")
    print()


def cmd_stats(args):
    """Show database statistics."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'build' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))

    print(f"\n{'='*60}")
    print(f"  Procurement Database — Statistics")
    print(f"{'='*60}")

    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    print(f"\n  File: {DB_PATH}")
    print(f"  Size: {db_size:.1f} MB")

    # Contracts
    c_count = conn.execute("SELECT COUNT(*) FROM contratos").fetchone()[0]
    c_with_nif = conn.execute("SELECT COUNT(*) FROM contratos WHERE adjudicante_nif != ''").fetchone()[0]
    c_with_price = conn.execute("SELECT COUNT(*) FROM contratos WHERE precoContratual > 0").fetchone()[0]
    c_with_link = conn.execute("SELECT COUNT(*) FROM contratos WHERE linkPecasProc != ''").fetchone()[0]
    total_value = conn.execute("SELECT SUM(precoContratual) FROM contratos WHERE precoContratual > 0").fetchone()[0] or 0

    print(f"\n  Contratos: {c_count:,} rows")
    print(f"    With NIF extracted: {c_with_nif:,} ({c_with_nif*100/c_count:.1f}%)")
    print(f"    With price: {c_with_price:,} ({c_with_price*100/c_count:.1f}%)")
    print(f"    With linkPecasProc: {c_with_link:,} ({c_with_link*100/c_count:.1f}%)")
    print(f"    Total value: €{total_value:,.0f}")

    # Top contract types
    print(f"\n  Top Contract Types:")
    for tipo, count in conn.execute(
        "SELECT tipoContrato, COUNT(*) FROM contratos WHERE tipoContrato != '' "
        "GROUP BY tipoContrato ORDER BY COUNT(*) DESC LIMIT 5"
    ).fetchall():
        print(f"    {tipo[:45]:45s} {count:>7,}")

    # Entities
    e_count = conn.execute("SELECT COUNT(*) FROM entidades").fetchone()[0]
    e_pt = conn.execute("SELECT COUNT(*) FROM entidades WHERE descPais = 'Portugal'").fetchone()[0]
    e_with_contracts = conn.execute("SELECT COUNT(*) FROM entidades WHERE numContratos > 0").fetchone()[0]

    print(f"\n  Entidades: {e_count:,} rows")
    print(f"    Portuguese: {e_pt:,} ({e_pt*100/e_count:.1f}%)")
    print(f"    With contracts: {e_with_contracts:,}")

    # Top entities by contract value
    print(f"\n  Top 5 Entities by Contract Value:")
    for name, val in conn.execute(
        "SELECT desigEntidade, totValorContratIni FROM entidades "
        "WHERE totValorContratIni > 0 ORDER BY totValorContratIni DESC LIMIT 5"
    ).fetchall():
        print(f"    {name[:45]:45s} €{val:>14,.0f}")

    print(f"\n{'='*60}\n")
    conn.close()


def cmd_query(args):
    """Run arbitrary SQL query."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'build' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    try:
        rows = conn.execute(args.sql).fetchall()
        if rows:
            # Print headers
            headers = rows[0].keys()
            widths = [max(len(str(h)), max(len(str(r[h])) for r in rows[:50])) for h in headers]
            header_line = " | ".join(f"{h:<{w}}" for h, w in zip(headers, widths))
            print(header_line)
            print("-" * len(header_line))
            for row in rows[:100]:
                print(" | ".join(f"{str(row[h]):<{w}}" for h, w in zip(headers, widths)))
            if len(rows) > 100:
                print(f"... ({len(rows) - 100} more rows)")
        else:
            print("  (no results)")
    except Exception as e:
        print(f"  SQL Error: {e}")

    conn.close()


def main():
    parser = argparse.ArgumentParser(
        description="Unified Procurement Database — SQLite for all procurement data",
    )
    sub = parser.add_subparsers(dest="command")

    dl = sub.add_parser("download", help="Download XLSX from dados.gov.pt")
    dl.add_argument("--years", type=int, nargs="+",
                    help="Specific years to download (default: all available)")
    dl.add_argument("--contratos-only", action="store_true",
                    help="Download only contracts XLSX")
    dl.add_argument("--entidades-only", action="store_true",
                    help="Download only entities XLSX")

    build = sub.add_parser("build", help="Build/refresh the database")
    build.add_argument("--force", action="store_true", help="Clear and rebuild")
    build.add_argument("--auto-download", action="store_true",
                       help="Download XLSX files if missing")

    sub.add_parser("status", help="Quick one-glance status overview")
    sub.add_parser("stats", help="Show database statistics")

    query = sub.add_parser("query", help="Run arbitrary SQL")
    query.add_argument("sql", help="SQL query to execute")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "download": cmd_download,
        "build": cmd_build,
        "status": cmd_status,
        "stats": cmd_stats,
        "query": cmd_query,
    }
    commands[args.command](args)


if __name__ == "__main__":
    main()
