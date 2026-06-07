#!/usr/bin/env python3
"""Announce Index — Download, parse, and analyze Portuguese tender announcements.

Downloads all 15 years (2012–2026) of anúncios XLSX files from dados.gov.pt,
indexes them in SQLite, and provides trend analysis, competition metrics,
and cross-reference with signed contract data.

Usage:
    python announce_index.py download        # Download all XLSX files
    python announce_index.py index           # Parse XLSX → SQLite
    python announce_index.py stats           # Summary statistics
    python announce_index.py trends          # Year-over-year trend analysis
    python announce_index.py sectors         # Breakdown by CPV/contract type
    python announce_index.py competition     # Competition quality metrics
    python announce_index.py entity --nif X  # Entity-specific announcements
    python announce_index.py search --query X # Full-text search descriptions
    python announce_index.py crossref        # Cross-ref with contract_index.json
    python announce_index.py export --out X  # Export to JSON
"""

import sys
import json
import re
import sqlite3
import argparse
import time
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone

try:
    import urllib.request
    import ssl
except ImportError:
    print("ERROR: urllib required (built-in)")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
DB_PATH = DATA_DIR / "anuncios_index.db"
CONTRACT_INDEX = DATA_DIR / "contract_index.json"
XLSX_DIR = DATA_DIR

# SSL context
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/octet-stream,*/*",
}

# Dataset ID on dados.gov.pt
DATASET_ID = "66d72fbc58cd7a63dae28712"
API_BASE = "https://dados.gov.pt/api/1/datasets/"


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def init_db() -> sqlite3.Connection:
    """Initialize the announcements database."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS anuncios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nAnuncio TEXT,
            IdIncm TEXT,
            dataPublicacao TEXT,
            nifEntidade TEXT,
            designacaoEntidade TEXT,
            descricaoAnuncio TEXT,
            url TEXT,
            numDR TEXT,
            serie TEXT,
            tipoActo TEXT,
            tiposContrato TEXT,
            PrecoBase REAL,
            CPVs TEXT,
            modeloAnuncio TEXT,
            Ano INTEGER,
            CriterAmbient TEXT,
            PrazoPropostas TEXT,
            PecasProcedimento TEXT,
            DataLimitePropostas TEXT,
            Lotes TEXT,
            UNIQUE(nAnuncio, nifEntidade, Ano)
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS download_log (
            year INTEGER PRIMARY KEY,
            filename TEXT,
            size_bytes INTEGER,
            rows INTEGER,
            downloaded_at TEXT,
            indexed_at TEXT
        )
    """)

    # Indexes for common queries
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ano ON anuncios(Ano)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_nif ON anuncios(nifEntidade)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tipo ON anuncios(tiposContrato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cpvs ON anuncios(CPVs)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_preco ON anuncios(PrecoBase)")

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def get_download_urls() -> dict[int, str]:
    """Fetch per-year download URLs from the dados.gov.pt API."""
    url = f"{API_BASE}{DATASET_ID}"
    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=15, context=SSL_CTX)
    data = json.loads(resp.read())

    urls = {}
    for r in data.get("resources", []):
        name = r.get("title", r.get("name", ""))
        fmt = r.get("format", "")
        dl_url = r.get("url", "")
        if fmt.upper() == "XLSX" and "anuncios" in name.lower():
            m = re.search(r"anuncios(\d{4})", name)
            if m:
                urls[int(m.group(1))] = dl_url

    return urls


def cmd_download(args):
    """Download all anúncios XLSX files."""
    print("Fetching download URLs from dados.gov.pt API...")
    urls = get_download_urls()
    print(f"  Found {len(urls)} XLSX files (years {min(urls)}-{max(urls)})\n")

    conn = init_db()
    total_size = 0

    for year in sorted(urls.keys()):
        xlsx_path = XLSX_DIR / f"anuncios{year}.xlsx"

        # Skip if already downloaded and >1KB
        if xlsx_path.exists() and xlsx_path.stat().st_size > 1000:
            size = xlsx_path.stat().st_size
            print(f"  {year}: already exists ({size:,} bytes)")
            total_size += size
            continue

        print(f"  {year}: downloading...", end=" ", flush=True)
        try:
            req = urllib.request.Request(urls[year], headers=HEADERS)
            resp = urllib.request.urlopen(req, timeout=60, context=SSL_CTX)
            data = resp.read()
            xlsx_path.write_bytes(data)
            size = len(data)
            total_size += size
            print(f"done ({size:,} bytes)")
        except Exception as e:
            print(f"FAILED: {e}")
            continue

        time.sleep(0.5)

    print(f"\n  Total: {total_size:,} bytes across {len(urls)} files")

    # Update download log
    for year, url in urls.items():
        xlsx_path = XLSX_DIR / f"anuncios{year}.xlsx"
        if xlsx_path.exists():
            conn.execute(
                "INSERT OR REPLACE INTO download_log (year, filename, size_bytes, downloaded_at) "
                "VALUES (?, ?, ?, ?)",
                (year, xlsx_path.name, xlsx_path.stat().st_size,
                 datetime.now(timezone.utc).isoformat()),
            )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Index (XLSX → SQLite)
# ---------------------------------------------------------------------------

def parse_xlsx(year: int, verbose: bool = False) -> list[tuple]:
    """Parse a single year's XLSX file into rows."""
    xlsx_path = XLSX_DIR / f"anuncios{year}.xlsx"
    if not xlsx_path.exists():
        if verbose:
            print(f"    {year}: file not found")
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        def get(field, default=""):
            idx = header_map.get(field)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return val if val is not None else default

        def get_num(field):
            val = get(field, None)
            if val is None:
                return None
            try:
                v = float(val)
                # Cap at €10 billion — values above are data entry errors
                if v > 1e10:
                    return None
                return v if v > 0 else None
            except (ValueError, TypeError):
                return None

        def fmt_date(val):
            if val is None:
                return ""
            if hasattr(val, 'strftime'):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()[:10]

        rows.append((
            str(get("nAnuncio", "")).strip(),
            str(get("IdIncm", "")).strip(),
            fmt_date(get("dataPublicacao")),
            str(get("nifEntidade", "")).strip(),
            str(get("designacaoEntidade", "")).strip(),
            str(get("descricaoAnuncio", "")).strip()[:2000],
            str(get("url", "")).strip(),
            str(get("numDR", "")).strip(),
            str(get("serie", "")).strip(),
            str(get("tipoActo", "")).strip(),
            str(get("tiposContrato", "")).strip(),
            get_num("PrecoBase"),
            str(get("CPVs", "")).strip(),
            str(get("modeloAnuncio", "")).strip(),
            year,
            str(get("CriterAmbient", "")).strip(),
            str(get("PrazoPropostas", "")).strip(),
            str(get("PecasProcedimento", "")).strip(),
            fmt_date(get("DataLimitePropostas")),
            str(get("Lotes", "")).strip(),
        ))

    wb.close()
    return rows


def cmd_index(args):
    """Parse all downloaded XLSX files into SQLite."""
    conn = init_db()

    # Clear existing data for fresh index
    if args.force:
        conn.execute("DELETE FROM anuncios")
        conn.commit()
        print("Cleared existing index.\n")

    years = range(2012, 2027)
    total_rows = 0
    total_new = 0

    for year in years:
        xlsx_path = XLSX_DIR / f"anuncios{year}.xlsx"
        if not xlsx_path.exists():
            print(f"  {year}: file not found, skipping")
            continue

        print(f"  {year}: parsing...", end=" ", flush=True)
        rows = parse_xlsx(year, verbose=args.verbose)

        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO anuncios "
            "(nAnuncio, IdIncm, dataPublicacao, nifEntidade, designacaoEntidade, "
            "descricaoAnuncio, url, numDR, serie, tipoActo, tiposContrato, "
            "PrecoBase, CPVs, modeloAnuncio, Ano, CriterAmbient, PrazoPropostas, "
            "PecasProcedimento, DataLimitePropostas, Lotes) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before

        total_new += new_count
        total_rows += len(rows)

        # Update download log
        conn.execute(
            "UPDATE download_log SET rows=?, indexed_at=? WHERE year=?",
            (len(rows), datetime.now(timezone.utc).isoformat(), year),
        )

        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    print(f"\n  Total: {total_rows:,} rows parsed, {total_new:,} inserted")
    print(f"  Database: {DB_PATH}")

    # Summary
    count = conn.execute("SELECT COUNT(*) FROM anuncios").fetchone()[0]
    print(f"  Index total: {count:,} announcements")

    conn.close()


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def cmd_stats(args):
    """Show summary statistics."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM anuncios").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    # Year distribution
    print(f"\n{'='*70}")
    print(f"  Anúncios Index — Statistics")
    print(f"{'='*70}")
    print(f"\n  Total announcements: {total:,}\n")

    print(f"  By Year:")
    year_data = conn.execute(
        "SELECT Ano, COUNT(*), SUM(CASE WHEN PrecoBase > 0 THEN PrecoBase ELSE 0 END) "
        "FROM anuncios GROUP BY Ano ORDER BY Ano"
    ).fetchall()
    for year, count, total_val in year_data:
        avg_val = (total_val / count) if count and total_val else 0
        print(f"    {year}: {count:>6,} announcements  "
              f"(total value: €{total_val:>15,.0f}, avg: €{avg_val:>12,.0f})")

    # Top entity types
    print(f"\n  Top Contract Types:")
    type_data = conn.execute(
        "SELECT tiposContrato, COUNT(*), SUM(COALESCE(PrecoBase, 0)) "
        "FROM anuncios WHERE tiposContrato != '' "
        "GROUP BY tiposContrato ORDER BY COUNT(*) DESC LIMIT 10"
    ).fetchall()
    for tipo, count, total_val in type_data:
        print(f"    {tipo[:50]:50s} {count:>6,} (€{total_val:>14,.0f})")

    # Top entities by announcement count
    print(f"\n  Top Entities (by announcement count):")
    entity_data = conn.execute(
        "SELECT designacaoEntidade, nifEntidade, COUNT(*), SUM(COALESCE(PrecoBase, 0)) "
        "FROM anuncios WHERE designacaoEntidade != '' "
        "GROUP BY nifEntidade ORDER BY COUNT(*) DESC LIMIT 10"
    ).fetchall()
    for name, nif, count, total_val in entity_data:
        print(f"    {name[:45]:45s} NIF={nif:>10s}  {count:>4,} (€{total_val:>12,.0f})")

    # Price statistics
    price_stats = conn.execute(
        "SELECT COUNT(*), MIN(PrecoBase), AVG(PrecoBase), MAX(PrecoBase), "
        "SUM(CASE WHEN PrecoBase > 0 THEN 1 ELSE 0 END) "
        "FROM anuncios"
    ).fetchone()
    print(f"\n  Price Statistics:")
    print(f"    Total records: {price_stats[0]:,}")
    print(f"    With price: {price_stats[4]:,} ({price_stats[4]*100/price_stats[0]:.1f}%)")
    if price_stats[4]:
        print(f"    Min: €{price_stats[1]:,.2f}")
        print(f"    Avg: €{price_stats[2]:,.2f}")
        print(f"    Max: €{price_stats[3]:,.2f}")

    print(f"{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Trends
# ---------------------------------------------------------------------------

def cmd_trends(args):
    """Year-over-year trend analysis."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM anuncios").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Anúncios — Year-over-Year Trends")
    print(f"{'='*70}\n")

    # Annual trends
    year_data = conn.execute(
        "SELECT Ano, COUNT(*) as cnt, "
        "SUM(COALESCE(PrecoBase, 0)) as total_val, "
        "AVG(CASE WHEN PrecoBase > 0 THEN PrecoBase END) as avg_val, "
        "COUNT(DISTINCT nifEntidade) as entities, "
        "COUNT(CASE WHEN tipoActo LIKE '%contrato%' THEN 1 END) as contratos, "
        "COUNT(CASE WHEN tipoActo LIKE '%procedimento%' THEN 1 END) as procedimentos "
        "FROM anuncios GROUP BY Ano ORDER BY Ano"
    ).fetchall()

    print(f"  {'Year':>6}  {'Count':>7}  {'Value (€)':>15}  {'Avg (€)':>12}  "
          f"{'Entities':>8}  {'YoY %':>7}")
    print(f"  {'─'*6}  {'─'*7}  {'─'*15}  {'─'*12}  {'─'*8}  {'─'*7}")

    prev_count = None
    for year, count, total_val, avg_val, entities, contratos, procedimentos in year_data:
        yoy = ""
        if prev_count and prev_count > 0:
            pct = (count - prev_count) / prev_count * 100
            yoy = f"{pct:+.1f}%"
        prev_count = count

        print(f"  {year:>6}  {count:>7,}  €{total_val:>14,.0f}  €{avg_val:>11,.0f}  "
              f"{entities:>8,}  {yoy:>7}")

    # Contract type evolution
    print(f"\n  Contract Type Evolution (top 5 types):")
    top_types = conn.execute(
        "SELECT tiposContrato, COUNT(*) FROM anuncios "
        "WHERE tiposContrato != '' GROUP BY tiposContrato ORDER BY COUNT(*) DESC LIMIT 5"
    ).fetchall()

    for tipo, _ in top_types:
        type_year = conn.execute(
            "SELECT Ano, COUNT(*) FROM anuncios WHERE tiposContrato = ? GROUP BY Ano ORDER BY Ano",
            (tipo,),
        ).fetchall()
        counts = {y: c for y, c in type_year}
        trend = " ".join(f"{y}:{counts.get(y, 0):>5}" for y, _, *_ in year_data[:5])
        print(f"    {tipo[:40]:40s}")

    # Publication day-of-week pattern
    print(f"\n  Publication Day Pattern:")
    dow_data = conn.execute(
        "SELECT CASE CAST(strftime('%w', dataPublicacao) AS INTEGER) "
        "WHEN 0 THEN 'Sun' WHEN 1 THEN 'Mon' WHEN 2 THEN 'Tue' "
        "WHEN 3 THEN 'Wed' WHEN 4 THEN 'Thu' WHEN 5 THEN 'Fri' "
        "WHEN 6 THEN 'Sat' ELSE '???' END as day, "
        "COUNT(*) FROM anuncios WHERE dataPublicacao != '' "
        "GROUP BY day ORDER BY COUNT(*) DESC"
    ).fetchall()
    for day, count in dow_data:
        bar = "█" * int(count / max(c for _, c in dow_data) * 30)
        print(f"    {day:>3}: {count:>6,} {bar}")

    # Monthly seasonality
    print(f"\n  Monthly Seasonality:")
    month_data = conn.execute(
        "SELECT CAST(strftime('%m', dataPublicacao) AS INTEGER) as m, COUNT(*) "
        "FROM anuncios WHERE dataPublicacao != '' GROUP BY m ORDER BY m"
    ).fetchall()
    max_cnt = max(c for _, c in month_data) if month_data else 1
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m, count in month_data:
        name = month_names[m - 1] if 1 <= m <= 12 else "???"
        bar = "█" * int(count / max_cnt * 30)
        print(f"    {name}: {count:>6,} {bar}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Sectors (CPV breakdown)
# ---------------------------------------------------------------------------

def cmd_sectors(args):
    """Breakdown by CPV code and contract type."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM anuncios").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Anúncios — Sector Breakdown")
    print(f"{'='*70}\n")

    # CPV code breakdown
    print(f"  Top CPV Codes (by count):")
    cpv_data = conn.execute(
        "SELECT CPVs, COUNT(*), SUM(COALESCE(PrecoBase, 0)) "
        "FROM anuncios WHERE CPVs != '' "
        "GROUP BY CPVs ORDER BY COUNT(*) DESC LIMIT 20"
    ).fetchall()
    for cpvs, count, total_val in cpv_data:
        # Extract first CPV code
        first_cpv = cpvs.split(";")[0].strip()[:50] if cpvs else ""
        print(f"    {first_cpv:50s} {count:>6,} (€{total_val:>14,.0f})")

    # Contract type breakdown
    print(f"\n  Contract Type Breakdown:")
    type_data = conn.execute(
        "SELECT tiposContrato, COUNT(*), SUM(COALESCE(PrecoBase, 0)), "
        "AVG(CASE WHEN PrecoBase > 0 THEN PrecoBase END) "
        "FROM anuncios WHERE tiposContrato != '' "
        "GROUP BY tiposContrato ORDER BY COUNT(*) DESC"
    ).fetchall()
    for tipo, count, total_val, avg_val in type_data:
        avg_str = f"€{avg_val:,.0f}" if avg_val else "N/A"
        print(f"    {tipo[:50]:50s} {count:>6,}  total €{total_val:>14,.0f}  avg {avg_str}")

    # Announcement type (modelo)
    print(f"\n  Announcement Model (modeloAnuncio):")
    model_data = conn.execute(
        "SELECT modeloAnuncio, COUNT(*) FROM anuncios "
        "WHERE modeloAnuncio != '' GROUP BY modeloAnuncio ORDER BY COUNT(*) DESC"
    ).fetchall()
    for model, count in model_data:
        pct = count * 100 / total
        bar = "█" * int(pct / 2)
        print(f"    {model[:50]:50s} {count:>6,} ({pct:>5.1f}%) {bar}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Competition
# ---------------------------------------------------------------------------

def cmd_competition(args):
    """Analyze competition quality metrics."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM anuncios").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Anúncios — Competition Quality Metrics")
    print(f"{'='*70}\n")

    print(f"  Announcement Act Types:")
    act_data = conn.execute(
        "SELECT tipoActo, COUNT(*), SUM(COALESCE(PrecoBase, 0)) "
        "FROM anuncios WHERE tipoActo != '' "
        "GROUP BY tipoActo ORDER BY COUNT(*) DESC LIMIT 15"
    ).fetchall()
    for acto, count, total_val in act_data:
        pct = count * 100 / total
        print(f"    {acto[:55]:55s} {count:>6,} ({pct:>5.1f}%)")

    print(f"\n  Direct Award vs Competitive:")
    ad_count = conn.execute(
        "SELECT COUNT(*) FROM anuncios WHERE tipoActo LIKE '%Ajuste Direto%'"
    ).fetchone()[0]
    comp_count = conn.execute(
        "SELECT COUNT(*) FROM anuncios WHERE tipoActo LIKE '%procedimento%' "
        "AND tipoActo NOT LIKE '%Ajuste Direto%'"
    ).fetchone()[0]
    other_count = total - ad_count - comp_count

    print(f"    Direct Awards (Ajuste Direto): {ad_count:>6,} ({ad_count*100/total:.1f}%)")
    print(f"    Competitive Procedures:        {comp_count:>6,} ({comp_count*100/total:.1f}%)")
    print(f"    Other/Unknown:                 {other_count:>6,} ({other_count*100/total:.1f}%)")

    print(f"\n  Bid Period Analysis:")
    short_deadline = conn.execute(
        "SELECT COUNT(*) FROM anuncios WHERE PrazoPropostas != '' "
        "AND PrazoPropostas GLOB '[0-9]*' AND CAST(PrazoPropostas AS INTEGER) < 15"
    ).fetchone()[0]
    medium_deadline = conn.execute(
        "SELECT COUNT(*) FROM anuncios WHERE PrazoPropostas != '' "
        "AND PrazoPropostas GLOB '[0-9]*' AND CAST(PrazoPropostas AS INTEGER) BETWEEN 15 AND 30"
    ).fetchone()[0]
    long_deadline = conn.execute(
        "SELECT COUNT(*) FROM anuncios WHERE PrazoPropostas != '' "
        "AND PrazoPropostas GLOB '[0-9]*' AND CAST(PrazoPropostas AS INTEGER) > 30"
    ).fetchone()[0]
    total_with_prazo = short_deadline + medium_deadline + long_deadline

    if total_with_prazo:
        print(f"    Short (<15 days):  {short_deadline:>6,} ({short_deadline*100/total_with_prazo:.1f}%)")
        print(f"    Medium (15-30 d):  {medium_deadline:>6,} ({medium_deadline*100/total_with_prazo:.1f}%)")
        print(f"    Long (>30 days):   {long_deadline:>6,} ({long_deadline*100/total_with_prazo:.1f}%)")

    # High-value direct awards
    print(f"\n  High-Value Direct Awards (>€1M):")
    high_ad = conn.execute(
        "SELECT designacaoEntidade, nifEntidade, PrecoBase, tiposContrato, dataPublicacao "
        "FROM anuncios WHERE tipoActo LIKE '%Ajuste Direto%' AND PrecoBase > 1000000 "
        "ORDER BY PrecoBase DESC LIMIT 15"
    ).fetchall()
    for name, nif, preco, tipo, data in high_ad:
        print(f"    €{preco:>14,.0f}  {name[:40]:40s}  {tipo[:25]:25s}  {data}")

    # Repeat entities (same NIF winning many direct awards)
    print(f"\n  Repeat Direct Award Entities (top 15):")
    repeat_data = conn.execute(
        "SELECT designacaoEntidade, nifEntidade, COUNT(*) as cnt, "
        "SUM(COALESCE(PrecoBase, 0)) as total_val "
        "FROM anuncios WHERE tipoActo LIKE '%Ajuste Direto%' AND nifEntidade != '' "
        "GROUP BY nifEntidade HAVING cnt > 1 "
        "ORDER BY cnt DESC LIMIT 15"
    ).fetchall()
    for name, nif, count, total_val in repeat_data:
        print(f"    {count:>4} awards  €{total_val:>14,.0f}  {name[:40]:40s}  NIF={nif}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Entity search
# ---------------------------------------------------------------------------

def cmd_entity(args):
    """Show announcements for a specific entity."""
    conn = init_db()

    where = ""
    params = []
    if args.nif:
        where = "WHERE nifEntidade = ?"
        params = [args.nif]
    elif args.name:
        where = "WHERE designacaoEntidade LIKE ?"
        params = [f"%{args.name}%"]
    else:
        print("  Specify --nif or --name")
        conn.close()
        return

    rows = conn.execute(
        f"SELECT nAnuncio, dataPublicacao, designacaoEntidade, tiposContrato, "
        f"PrecoBase, tipoActo, descricaoAnuncio, CPVs "
        f"FROM anuncios {where} ORDER BY dataPublicacao DESC LIMIT ?",
        params + [args.limit],
    ).fetchall()

    if not rows:
        print("  No announcements found.")
        conn.close()
        return

    # Entity summary
    entity = conn.execute(
        f"SELECT designacaoEntidade, COUNT(*), SUM(COALESCE(PrecoBase, 0)) "
        f"FROM anuncios {where}",
        params,
    ).fetchone()

    print(f"\n{'='*70}")
    print(f"  Entity: {entity[0]}")
    print(f"  Total announcements: {entity[1]:,}")
    print(f"  Total value: €{entity[2]:,.0f}")
    print(f"{'='*70}\n")

    for nAnuncio, data, name, tipo, preco, acto, desc, cpvs in rows:
        preco_str = f"€{preco:,.0f}" if preco else "N/A"
        print(f"  [{data}] {nAnuncio}")
        print(f"    Type: {tipo} | Act: {acto} | Value: {preco_str}")
        if desc:
            print(f"    {desc[:80]}")
        print()

    conn.close()


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

def cmd_search(args):
    """Full-text search in announcement descriptions."""
    conn = init_db()

    query = f"%{args.query}%"
    rows = conn.execute(
        "SELECT nAnuncio, dataPublicacao, designacaoEntidade, tiposContrato, "
        "PrecoBase, tipoActo, descricaoAnuncio "
        "FROM anuncios WHERE descricaoAnuncio LIKE ? OR designacaoEntidade LIKE ? "
        "ORDER BY PrecoBase DESC LIMIT ?",
        (query, query, args.limit),
    ).fetchall()

    if not rows:
        print("  No results found.")
        conn.close()
        return

    print(f"\n  Found {len(rows)} results for '{args.query}':\n")
    for nAnuncio, data, name, tipo, preco, acto, desc in rows:
        preco_str = f"€{preco:,.0f}" if preco else "N/A"
        print(f"  [{data}] {nAnuncio} — {name[:45]}")
        print(f"    {tipo} | {acto} | {preco_str}")
        if desc:
            print(f"    {desc[:80]}")
        print()

    conn.close()


# ---------------------------------------------------------------------------
# Cross-reference with contracts
# ---------------------------------------------------------------------------

def cmd_crossref(args):
    """Cross-reference announcements with signed contracts."""
    conn = init_db()

    if not CONTRACT_INDEX.exists():
        print(f"  Contract index not found at {CONTRACT_INDEX}")
        print("  Run bep_base_crossref.py first.")
        conn.close()
        return

    with open(CONTRACT_INDEX, "r", encoding="utf-8") as f:
        contract_index = json.load(f)

    # Build NIF → contracts lookup
    nif_contracts = defaultdict(list)
    for nif, contracts in contract_index.items():
        for c in contracts:
            nif_contracts[nif].append(c)

    print(f"\n{'='*70}")
    print(f"  Anúncios × Contracts Cross-Reference")
    print(f"{'='*70}")
    print(f"  Announcements in index: {conn.execute('SELECT COUNT(*) FROM anuncios').fetchone()[0]:,}")
    print(f"  Contracts in index: {sum(len(v) for v in contract_index.values()):,}")
    print(f"  Entities with contracts: {len(contract_index):,}")

    anuncio_nifs = set(r[0] for r in conn.execute(
        "SELECT DISTINCT nifEntidade FROM anuncios WHERE nifEntidade != ''"
    ).fetchall())
    contract_nifs = set(contract_index.keys())
    both = anuncio_nifs & contract_nifs

    print(f"  Entities in BOTH: {len(both):,}")
    print(f"  Entities only in announcements: {len(anuncio_nifs - contract_nifs):,}")
    print(f"  Entities only in contracts: {len(contract_nifs - anuncio_nifs):,}")

    print(f"\n  Potential Matches (announcement → contract):")
    matches = []
    for nif in sorted(both):
        a_count = conn.execute(
            "SELECT COUNT(*) FROM anuncios WHERE nifEntidade = ? AND PrecoBase > 0",
            (nif,),
        ).fetchone()[0]
        c_count = len(nif_contracts.get(nif, []))

        entity_name = conn.execute(
            "SELECT designacaoEntidade FROM anuncios WHERE nifEntidade = ? LIMIT 1",
            (nif,),
        ).fetchone()
        name = entity_name[0] if entity_name else nif

        if a_count and c_count:
            matches.append((name, nif, a_count, c_count))

    matches.sort(key=lambda x: -(x[2] + x[3]))
    for name, nif, a_count, c_count in matches[:20]:
        print(f"    {name[:45]:45s} NIF={nif:>10s}  {a_count} announcements, {c_count} contracts")

    print(f"{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def cmd_export(args):
    """Export index to JSON."""
    conn = init_db()

    rows = conn.execute(
        "SELECT nAnuncio, dataPublicacao, nifEntidade, designacaoEntidade, "
        "tiposContrato, PrecoBase, tipoActo, CPVs, Ano, url "
        "FROM anuncios ORDER BY Ano, dataPublicacao"
    ).fetchall()

    data = []
    for row in rows:
        data.append({
            "nAnuncio": row[0],
            "dataPublicacao": row[1],
            "nifEntidade": row[2],
            "designacaoEntidade": row[3],
            "tiposContrato": row[4],
            "precoBase": row[5],
            "tipoActo": row[6],
            "cpvs": row[7],
            "ano": row[8],
            "url": row[9],
        })

    out_path = Path(args.out) if args.out else DATA_DIR / "anuncios_index.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"  Exported {len(data):,} announcements to {out_path}")
    conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Announce Index — Portuguese tender announcements (2012–2026)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    sub = parser.add_subparsers(dest="command")

    # Download
    sub.add_parser("download", help="Download all XLSX files from dados.gov.pt")

    # Index
    idx = sub.add_parser("index", help="Parse XLSX files into SQLite index")
    idx.add_argument("--force", action="store_true", help="Re-index from scratch")
    idx.add_argument("-v", "--verbose", action="store_true")

    # Stats
    sub.add_parser("stats", help="Summary statistics")

    # Trends
    sub.add_parser("trends", help="Year-over-year trend analysis")

    # Sectors
    sub.add_parser("sectors", help="Breakdown by CPV/contract type")

    # Competition
    sub.add_parser("competition", help="Competition quality metrics")

    # Entity
    ent = sub.add_parser("entity", help="Entity-specific announcements")
    ent.add_argument("--nif", help="Filter by entity NIF")
    ent.add_argument("--name", help="Filter by entity name (partial match)")
    ent.add_argument("--limit", type=int, default=20)

    # Search
    srch = sub.add_parser("search", help="Full-text search descriptions")
    srch.add_argument("--query", required=True, help="Search query")
    srch.add_argument("--limit", type=int, default=20)

    # Crossref
    sub.add_parser("crossref", help="Cross-ref with contract_index.json")

    # Export
    exp = sub.add_parser("export", help="Export to JSON")
    exp.add_argument("--out", help="Output file path")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "download": cmd_download,
        "index": cmd_index,
        "stats": cmd_stats,
        "trends": cmd_trends,
        "sectors": cmd_sectors,
        "competition": cmd_competition,
        "entity": cmd_entity,
        "search": cmd_search,
        "crossref": cmd_crossref,
        "export": cmd_export,
    }

    commands[args.command](args)


if __name__ == "__main__":
    main()
