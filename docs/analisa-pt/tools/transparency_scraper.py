#!/usr/bin/env python3
"""Transparency Scraper — Download and analyze Portuguese budget execution and PRR data.

Downloads budget execution (execução orçamental) and PRR (Plano de Recuperação
e Resiliência) datasets from dados.gov.pt, indexes them in SQLite, and provides
analysis commands for corruption detection.

Usage:
    python transparency_scraper.py download              # Download all datasets
    python transparency_scraper.py download --type prr   # Download PRR only
    python transparency_scraper.py download --type budget # Download budget only
    python transparency_scraper.py index                 # Parse XLSX → SQLite
    python transparency_scraper.py stats                 # Summary statistics
    python transparency_scraper.py prr                   # PRR contract analysis
    python transparency_scraper.py budget                # Budget execution analysis
    python transparency_scraper.py crossref              # Cross-ref with procurement.db
    python transparency_scraper.py entity --nif X        # Entity-specific analysis
    python transparency_scraper.py search --query X      # Full-text search
    python transparency_scraper.py export --out X        # Export to JSON
"""

import sys
import json
import re
import sqlite3
import argparse
import time
from pathlib import Path
from datetime import datetime, timezone

from utils_db import connect as db_connect

try:
    import urllib.request
    import ssl
except ImportError:
    print("ERROR: urllib required (built-in)")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
DB_PATH = DATA_DIR / "transparency.db"
PROCUREMENT_DB = DATA_DIR / "procurement.db"
XLSX_DIR = DATA_DIR

# SSL context
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/octet-stream,*/*",
}

# dados.gov.pt dataset IDs
API_BASE = "https://dados.gov.pt/api/1/datasets/"

# Budget Execution datasets (by economic classification)
BUDGET_DATASETS = {
    "expense_economic": "60d9da99078190458358446c",   # Despesa por classificação económica
    "expense_functional": "60dde560078190da7dc26d5c",  # Despesa por classificação funcional
    "revenue": "60d9db7507819049478cf8ac",              # Receita por classificação económica
    "central_indicators": "60d9dbe90781904b6e3c4a31",  # Indicadores Administração Central
}

# PRR datasets
PRR_DATASETS = {
    "contracts": "6717db451b0eaad60ac5008c",           # Contratos Públicos PRR
    "entities": "6717dca11b0eaad60ac5008f",             # Entidades
    "projects": "6717dcff1b0eaad60ac50090",             # Projetos
    "locations": "6717dc1f1b0eaad60ac5008e",            # Localizações
    "milestones": "6718cbd7d7c1ce7589c50090",           # Marcos e Metas
    "investment": "6718cb20d7c1ce7589c5008c",           # Investimentos
    "criteria": "6717dbd71b0eaad60ac5008d",             # Critérios de Atribuição
    "entity_contracts": "6718c29ebbf3654d2bc5008c",    # Entidades - Contratos Públicos
}


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def init_db() -> sqlite3.Connection:
    """Initialize the transparency database."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    # PRAGMA-tuned connect via utils_db.connect (WAL, 200MB cache, mmap on POSIX).
    conn = db_connect(str(DB_PATH))

    # PRR contracts table (from prr_contracts_*.xlsx)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cd_contrato TEXT,
            dt_referencia TEXT,
            ds_contrato TEXT,
            sumario TEXT,
            cd_base_gov TEXT,
            dt_assinatura TEXT,
            montante REAL,
            cd_projeto TEXT,
            ds_projeto TEXT,
            perc_projeto REAL,
            UNIQUE(cd_contrato)
        )
    """)

    # PRR entity-contract links (from prr_entity_contracts_*.xlsx)
    # This links entities to contracts (buyer-supplier relationship)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_entity_contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cd_contrato TEXT,
            dt_referencia TEXT,
            ds_contrato TEXT,
            cd_entidade TEXT,
            ds_entidade TEXT,
            ds_papel TEXT,
            valor_contrato REAL,
            UNIQUE(cd_contrato, cd_entidade, ds_papel)
        )
    """)

    # PRR entities (from prr_entities_*.xlsx)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_entities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cd_entidade TEXT,
            dt_referencia TEXT,
            nif TEXT,
            ds_entidade TEXT,
            papel TEXT,
            atividade_economica TEXT,
            localizacao TEXT,
            valor_contratado REAL,
            valor_pago REAL,
            cd_projeto TEXT,
            UNIQUE(cd_entidade)
        )
    """)

    # PRR projects (from prr_projects_*.xlsx)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cd_projeto TEXT,
            dt_referencia TEXT,
            ds_projeto TEXT,
            sumario TEXT,
            valor_aprovado REAL,
            valor_pago REAL,
            subvencoes REAL,
            emprestimos REAL,
            nota_final REAL,
            cd_investimento TEXT,
            dt_inicio TEXT,
            dt_prevista_conclusao TEXT,
            dt_efetiva_conclusao TEXT,
            UNIQUE(cd_projeto)
        )
    """)

    # PRR locations (from prr_locations_*.xlsx)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_locations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cd_projeto TEXT,
            dt_referencia TEXT,
            cd_nutsii TEXT,
            ds_nutsii TEXT,
            cd_nutsiii TEXT,
            ds_nutsiii TEXT,
            cd_distrito TEXT,
            ds_distrito TEXT,
            cd_concelho TEXT,
            ds_concelho TEXT,
            perc_valor_aprovado REAL,
            perc_valor_pago REAL,
            UNIQUE(cd_projeto, cd_concelho)
        )
    """)

    # PRR milestones (from prr_milestones_*.xlsx)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prr_milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            componente TEXT,
            data_ref TEXT,
            sequencial INTEGER,
            codigo_reforma TEXT,
            designacao_reforma TEXT,
            tipo TEXT,
            designacao TEXT,
            indicador_qualitativo TEXT,
            indicador_quantitativo TEXT,
            referencia TEXT,
            objetivo TEXT,
            trimestre TEXT,
            ano INTEGER,
            fonte_dados TEXT,
            responsabilidade TEXT,
            descricao TEXT,
            pressupostos_riscos TEXT,
            mecanismo_verificacao TEXT,
            indicadores_desembolso TEXT,
            natureza_medida TEXT,
            data_conclusao TEXT,
            valor_atingido REAL
        )
    """)

    # Budget execution table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS budget (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_key TEXT,
            ano INTEGER,
            mes INTEGER,
            nivel_orcamental TEXT,
            descricao TEXT,
            valor_previsto REAL,
            valor_realizado REAL,
            percentagem REAL,
            UNIQUE(dataset_key, ano, mes, nivel_orcamental, descricao)
        )
    """)

    # Download log
    conn.execute("""
        CREATE TABLE IF NOT EXISTS download_log (
            key TEXT PRIMARY KEY,
            dataset_id TEXT,
            filename TEXT,
            size_bytes INTEGER,
            rows INTEGER,
            downloaded_at TEXT,
            indexed_at TEXT
        )
    """)

    # Indexes
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_c_contrato ON prr_contracts(cd_contrato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_c_projeto ON prr_contracts(cd_projeto)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_c_base ON prr_contracts(cd_base_gov)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_ec_entidade ON prr_entity_contracts(cd_entidade)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_ec_contrato ON prr_entity_contracts(cd_contrato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_e_nif ON prr_entities(nif)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_e_papel ON prr_entities(papel)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_p_projeto ON prr_projects(cd_projeto)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_prr_l_concelho ON prr_locations(cd_concelho)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_budget_ano ON budget(ano)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_budget_desc ON budget(descricao)")

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# API Helpers
# ---------------------------------------------------------------------------

def api_get(dataset_id: str) -> dict:
    """Fetch dataset metadata from dados.gov.pt API."""
    url = f"{API_BASE}{dataset_id}"
    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=15, context=SSL_CTX)
    data = json.loads(resp.read())
    return data.get("data", data)


def get_resource_urls(dataset_id: str, fmt: str = "xlsx") -> list[dict]:
    """Get download URLs for a dataset's resources, filtered by format."""
    ds = api_get(dataset_id)
    resources = []
    for r in ds.get("resources", []):
        r_fmt = str(r.get("format", "")).lower()
        if r_fmt == fmt.lower():
            resources.append({
                "title": r.get("title", r.get("name", "")),
                "url": r.get("url", ""),
                "format": r_fmt,
            })
    return resources


def download_file(url: str, local_path: Path) -> bool:
    """Download a file from URL to local path."""
    if local_path.exists() and local_path.stat().st_size > 1000:
        return True  # Already downloaded

    try:
        req = urllib.request.Request(url, headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=60, context=SSL_CTX)
        data = resp.read()
        local_path.write_bytes(data)
        return True
    except Exception as e:
        print(f"    FAILED: {e}")
        return False


def _safe_workbook(path: Path):
    """Open an XLSX workbook with error handling. Returns (wb, ws) or (None, None)."""
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True)
        return wb, wb.active
    except Exception as e:
        print(f"    WARNING: Could not open {path.name}: {e}")
        return None, None


def _cell(row, idx, default=""):
    """Safely get a cell value from a row tuple."""
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return val if val is not None else default


def _cell_num(row, idx):
    """Get a numeric cell value, or None."""
    val = _cell(row, idx, None)
    if val is None:
        return None
    try:
        v = float(val)
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def _cell_int(row, idx, default=None):
    """Get an integer cell value."""
    val = _cell(row, idx, None)
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _fmt_date(val):
    """Format a date value to YYYY-MM-DD."""
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return s


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def cmd_download(args):
    """Download datasets from dados.gov.pt."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = init_db()

    download_type = args.type if hasattr(args, "type") and args.type else "all"

    total_size = 0
    total_files = 0

    # --- Budget Execution ---
    if download_type in ("all", "budget"):
        print("\n" + "=" * 70)
        print("  Budget Execution Datasets (Execução Orçamental)")
        print("=" * 70)

        for key, ds_id in BUDGET_DATASETS.items():
            print(f"\n  Dataset: {key}")
            try:
                resources = get_resource_urls(ds_id)
                print(f"    Found {len(resources)} XLSX files")
            except Exception as e:
                print(f"    ERROR fetching metadata: {e}")
                continue

            for r in resources:
                title = r["title"]
                url = r["url"]
                fname = f"budget_{key}_{title}.xlsx"
                local_path = XLSX_DIR / fname

                if local_path.exists() and local_path.stat().st_size > 1000:
                    print(f"    {title}: already exists")
                    continue

                print(f"    {title}: downloading...", end=" ", flush=True)
                if download_file(url, local_path):
                    size = local_path.stat().st_size
                    total_size += size
                    total_files += 1
                    print(f"done ({size:,} bytes)")

                    conn.execute(
                        "INSERT OR REPLACE INTO download_log "
                        "(key, dataset_id, filename, size_bytes, downloaded_at) "
                        "VALUES (?, ?, ?, ?, ?)",
                        (f"budget_{key}_{title}", ds_id, fname, size,
                         datetime.now(timezone.utc).isoformat()),
                    )
                time.sleep(0.3)

    # --- PRR ---
    if download_type in ("all", "prr"):
        print("\n" + "=" * 70)
        print("  PRR Datasets (Plano de Recuperação e Resiliência)")
        print("=" * 70)

        for key, ds_id in PRR_DATASETS.items():
            print(f"\n  Dataset: {key}")
            try:
                resources = get_resource_urls(ds_id)
                print(f"    Found {len(resources)} file(s)")
            except Exception as e:
                print(f"    ERROR fetching metadata: {e}")
                continue

            for r in resources:
                title = r["title"]
                url = r["url"]
                fname = f"prr_{key}_{title}"
                local_path = XLSX_DIR / fname

                if local_path.exists() and local_path.stat().st_size > 1000:
                    print(f"    {title}: already exists")
                    continue

                print(f"    {title}: downloading...", end=" ", flush=True)
                if download_file(url, local_path):
                    size = local_path.stat().st_size
                    total_size += size
                    total_files += 1
                    print(f"done ({size:,} bytes)")

                    conn.execute(
                        "INSERT OR REPLACE INTO download_log "
                        "(key, dataset_id, filename, size_bytes, downloaded_at) "
                        "VALUES (?, ?, ?, ?, ?)",
                        (f"prr_{key}_{title}", ds_id, fname, size,
                         datetime.now(timezone.utc).isoformat()),
                    )
                time.sleep(0.3)

    conn.commit()
    conn.close()

    print(f"\n  Total: {total_files} new files, {total_size:,} bytes")


# ---------------------------------------------------------------------------
# Index — PRR Contracts
# ---------------------------------------------------------------------------

def _parse_prr_contracts(xlsx_path: Path) -> list[tuple]:
    """Parse a PRR contracts XLSX file.

    Columns: cd_contrato, dt_referencia, ds_contrato, sumario,
             cd_base_gov, dt_assinatura_contrato, montante_contratualizado,
             cd_projeto, ds_projeto, perc_contrato_projeto
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cd = _cell(row, h.get("cd_contrato"))
            if not cd:
                continue

            rows.append((
                str(cd).strip(),
                _fmt_date(_cell(row, h.get("dt_referencia"))),
                str(_cell(row, h.get("ds_contrato"))).strip()[:500],
                str(_cell(row, h.get("sumario"))).strip()[:500],
                str(_cell(row, h.get("cd_base_gov"))).strip(),
                _fmt_date(_cell(row, h.get("dt_assinatura_contrato"))),
                _cell_num(row, h.get("montante_contratualizado")),
                str(_cell(row, h.get("cd_projeto"))).strip(),
                str(_cell(row, h.get("ds_projeto"))).strip()[:200],
                _cell_num(row, h.get("perc_contrato_projeto")),
            ))
    finally:
        wb.close()

    return rows


def _parse_prr_entity_contracts(xlsx_path: Path) -> list[tuple]:
    """Parse a PRR entity-contract links XLSX file.

    Columns: cd_contrato, dt_referencia, ds_contrato, cd_entidade,
             ds_entidade, ds_papel_entidade_contrato, valor_contrato
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cd = _cell(row, h.get("cd_contrato"))
            ent = _cell(row, h.get("cd_entidade"))
            if not cd:
                continue

            rows.append((
                str(cd).strip(),
                _fmt_date(_cell(row, h.get("dt_referencia"))),
                str(_cell(row, h.get("ds_contrato"))).strip()[:500],
                str(ent).strip() if ent else "",
                str(_cell(row, h.get("ds_entidade"))).strip()[:200],
                str(_cell(row, h.get("ds_papel_entidade_contrato"))).strip(),
                _cell_num(row, h.get("valor_contrato")),
            ))
    finally:
        wb.close()

    return rows


def _stream_prr_entities(xlsx_path: Path, conn: sqlite3.Connection, batch_size: int = 50000):
    """Stream-parse PRR entities XLSX directly into SQLite in batches.

    Avoids loading the entire file into memory. Returns total row count.
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return 0

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    total_rows = 0
    batch = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cd = _cell(row, h.get("cd_entidade"))
            if not cd:
                continue

            batch.append((
                str(cd).strip(),
                _fmt_date(_cell(row, h.get("dt_referencia"))),
                str(_cell(row, h.get("nif_entidade"))).strip(),
                str(_cell(row, h.get("ds_entidade"))).strip()[:200],
                str(_cell(row, h.get("papel_entidade"))).strip(),
                str(_cell(row, h.get("atividade_economica"))).strip(),
                str(_cell(row, h.get("localizacao_sede"))).strip(),
                _cell_num(row, h.get("valor_contratado")),
                _cell_num(row, h.get("valor_pago")),
                str(_cell(row, h.get("cd_projeto"))).strip(),
            ))

            if len(batch) >= batch_size:
                changes_before = conn.total_changes
                conn.executemany(
                    "INSERT OR IGNORE INTO prr_entities "
                    "(cd_entidade, dt_referencia, nif, ds_entidade, papel, "
                    "atividade_economica, localizacao, valor_contratado, valor_pago, cd_projeto) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?)",
                    batch,
                )
                total_rows += conn.total_changes - changes_before
                conn.commit()
                batch.clear()
    finally:
        # Flush remaining rows
        if batch:
            changes_before = conn.total_changes
            conn.executemany(
                "INSERT OR IGNORE INTO prr_entities "
                "(cd_entidade, dt_referencia, nif, ds_entidade, papel, "
                "atividade_economica, localizacao, valor_contratado, valor_pago, cd_projeto) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                batch,
            )
            total_rows += conn.total_changes - changes_before
            conn.commit()
        wb.close()

    return total_rows


def _parse_prr_projects(xlsx_path: Path) -> list[tuple]:
    """Parse a PRR projects XLSX file.

    Columns: cd_projeto, dt_referencia, ds_projeto, sumario,
             valor_aprovado, valor_pago, subvencoes, emprestimos,
             nota_final_candidatura, cd_investimento, dt_inicio,
             dt_prevista_conclusao, dt_efetiva_conclusao
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cd = _cell(row, h.get("cd_projeto"))
            if not cd:
                continue

            rows.append((
                str(cd).strip(),
                _fmt_date(_cell(row, h.get("dt_referencia"))),
                str(_cell(row, h.get("ds_projeto"))).strip()[:300],
                str(_cell(row, h.get("sumario"))).strip()[:500],
                _cell_num(row, h.get("valor_aprovado")),
                _cell_num(row, h.get("valor_pago")),
                _cell_num(row, h.get("subvencoes")),
                _cell_num(row, h.get("emprestimos")),
                _cell_num(row, h.get("nota_final_candidatura")),
                str(_cell(row, h.get("cd_investimento"))).strip(),
                _fmt_date(_cell(row, h.get("dt_inicio"))),
                _fmt_date(_cell(row, h.get("dt_prevista_conclusao"))),
                _fmt_date(_cell(row, h.get("dt_efetiva_conclusao"))),
            ))
    finally:
        wb.close()

    return rows


def _parse_prr_locations(xlsx_path: Path) -> list[tuple]:
    """Parse a PRR locations XLSX file.

    Columns: cd_projeto, dt_referencia, cd_nutsii, ds_nutsii,
             cd_nutsiii, ds_nutsiii, cd_distrito, ds_distrito,
             cd_concelho, ds_concelho, perc_valor_aprovado, perc_valor_pago
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cd = _cell(row, h.get("cd_projeto"))
            concelho = _cell(row, h.get("cd_concelho"))
            if not cd:
                continue

            rows.append((
                str(cd).strip(),
                _fmt_date(_cell(row, h.get("dt_referencia"))),
                str(_cell(row, h.get("cd_nutsii"))).strip(),
                str(_cell(row, h.get("ds_nutsii"))).strip(),
                str(_cell(row, h.get("cd_nutsiii"))).strip(),
                str(_cell(row, h.get("ds_nutsiii"))).strip(),
                str(_cell(row, h.get("cd_distrito"))).strip(),
                str(_cell(row, h.get("ds_distrito"))).strip(),
                str(concelho).strip() if concelho else "",
                str(_cell(row, h.get("ds_concelho"))).strip(),
                _cell_num(row, h.get("perc_valor_aprovado")),
                _cell_num(row, h.get("perc_valor_pago")),
            ))
    finally:
        wb.close()

    return rows


def _parse_prr_milestones(xlsx_path: Path) -> list[tuple]:
    """Parse a PRR milestones XLSX file.

    Columns: Componente, Data, Número Sequencial,
             Código da Reforma ou Investimento, Designação da Reforma ou Investimento,
             Marco / Meta, Designação, Indicador Qualitativo (para marco),
             Indicador Quantitativo (para meta), Referência, Objetivo,
             Trimestre, Ano, Fonte de dados/Metodologia, ...
    """
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name.strip(): i for i, name in enumerate(headers) if name}

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            componente = _cell(row, h.get("Componente"))
            if not componente:
                continue

            rows.append((
                str(componente).strip(),
                _fmt_date(_cell(row, h.get("Data"))),
                _cell_int(row, h.get("Número Sequencial")),
                str(_cell(row, h.get("Código da Reforma ou Investimento"))).strip(),
                str(_cell(row, h.get("Designação da Reforma ou Investimento"))).strip()[:200],
                str(_cell(row, h.get("Marco / Meta"))).strip(),
                str(_cell(row, h.get("Designação"))).strip()[:200],
                str(_cell(row, h.get("Indicador Qualitativo (para marco)"))).strip()[:200],
                str(_cell(row, h.get("Indicador Quantitativo (para meta)"))).strip()[:200],
                str(_cell(row, h.get("Referência"))).strip(),
                str(_cell(row, h.get("Objetivo"))).strip(),
                str(_cell(row, h.get("Trimestre"))).strip(),
                _cell_int(row, h.get("Ano")),
                str(_cell(row, h.get("Fonte de dados/Metodologia"))).strip()[:200],
                str(_cell(row, h.get("Responsabilidade pelo reporte e implementação"))).strip()[:200],
                str(_cell(row, h.get("Descrição/definição de cada marco e meta"))).strip()[:500],
                str(_cell(row, h.get("Pressupostos/Riscos"))).strip()[:200],
                str(_cell(row, h.get("Mecanismo de verificação"))).strip()[:200],
                str(_cell(row, h.get("Indicadores para desembolso"))).strip()[:200],
                str(_cell(row, h.get("Natureza da medida"))).strip(),
                _fmt_date(_cell(row, h.get("Data de Conclusão"))),
                _cell_num(row, h.get("Valor Atingido")),
            ))
    finally:
        wb.close()

    return rows


def _parse_budget(xlsx_path: Path, dataset_key: str) -> list[tuple]:
    """Parse a budget execution XLSX file into rows."""
    wb, ws = _safe_workbook(xlsx_path)
    if ws is None:
        return []

    headers = [cell.value for cell in ws[1]]
    h = {name: i for i, name in enumerate(headers) if name}

    # Try to extract year from filename
    fname = xlsx_path.name
    year_match = re.search(r"(\d{4})", fname)
    default_year = int(year_match.group(1)) if year_match else None

    rows = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            nivel_key = (h.get("Nível Orçamental") or h.get("Nível orçamental")
                         or h.get("Classificação") or h.get("Código"))
            nivel = str(_cell(row, nivel_key)).strip()
            desc_key = (h.get("Descrição") or h.get("Descrição da Despesa")
                        or h.get("Descrição da Receita"))
            desc = str(_cell(row, desc_key)).strip()
            if not nivel and not desc:
                continue

            ano = _cell_int(row, h.get("Ano", h.get("ano")), default_year)
            mes = _cell_int(row, h.get("Mês", h.get("mes")), 0)

            rows.append((
                dataset_key,
                ano,
                mes,
                nivel[:200],
                desc[:500],
                _cell_num(row, h.get("Valor Previsto",
                           h.get("Despesa prevista",
                           h.get("Receita prevista")))),
                _cell_num(row, h.get("Valor Realizado",
                           h.get("Despesa realizada",
                           h.get("Receita realizada")))),
                _cell_num(row, h.get("Percentagem",
                           h.get("% Execução",
                           h.get("%")))),
            ))
    finally:
        wb.close()

    return rows


# ---------------------------------------------------------------------------
# Index command
# ---------------------------------------------------------------------------

def cmd_index(args):
    """Parse all downloaded XLSX files into SQLite."""
    conn = init_db()

    if args.force:
        for table in ("prr_contracts", "prr_entity_contracts", "prr_entities",
                       "prr_projects", "prr_locations", "prr_milestones", "budget"):
            conn.execute(f"DELETE FROM {table}")
        conn.commit()
        print("Cleared existing index.\n")

    total_new = 0

    # --- Index PRR contracts ---
    print("Indexing PRR contracts...")
    for f in sorted(XLSX_DIR.glob("prr_contracts_*.xlsx")):
        print(f"  {f.name}: parsing...", end=" ", flush=True)
        rows = _parse_prr_contracts(f)
        if not rows:
            print("0 rows")
            continue
        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO prr_contracts "
            "(cd_contrato, dt_referencia, ds_contrato, sumario, cd_base_gov, "
            "dt_assinatura, montante, cd_projeto, ds_projeto, perc_projeto) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    # --- Index PRR entity-contract links ---
    print("\nIndexing PRR entity-contract links...")
    for f in sorted(XLSX_DIR.glob("prr_entity_contracts_*.xlsx")):
        print(f"  {f.name}: parsing...", end=" ", flush=True)
        rows = _parse_prr_entity_contracts(f)
        if not rows:
            print("0 rows")
            continue
        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO prr_entity_contracts "
            "(cd_contrato, dt_referencia, ds_contrato, cd_entidade, "
            "ds_entidade, ds_papel, valor_contrato) "
            "VALUES (?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    # --- Index PRR entities (streaming to avoid MemoryError) ---
    print("\nIndexing PRR entities...")
    for f in sorted(XLSX_DIR.glob("prr_entities_*.xlsx")):
        print(f"  {f.name}: streaming to DB in batches...", end=" ", flush=True)
        new_count = _stream_prr_entities(f, conn)
        total_new += new_count
        print(f"{new_count:,} new rows")
        conn.commit()

    # --- Index PRR projects ---
    print("\nIndexing PRR projects...")
    for f in sorted(XLSX_DIR.glob("prr_projects_*.xlsx")):
        print(f"  {f.name}: parsing...", end=" ", flush=True)
        rows = _parse_prr_projects(f)
        if not rows:
            print("0 rows")
            continue
        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO prr_projects "
            "(cd_projeto, dt_referencia, ds_projeto, sumario, valor_aprovado, "
            "valor_pago, subvencoes, emprestimos, nota_final, cd_investimento, "
            "dt_inicio, dt_prevista_conclusao, dt_efetiva_conclusao) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    # --- Index PRR locations ---
    print("\nIndexing PRR locations...")
    for f in sorted(XLSX_DIR.glob("prr_locations_*.xlsx")):
        print(f"  {f.name}: parsing...", end=" ", flush=True)
        rows = _parse_prr_locations(f)
        if not rows:
            print("0 rows")
            continue
        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO prr_locations "
            "(cd_projeto, dt_referencia, cd_nutsii, ds_nutsii, cd_nutsiii, "
            "ds_nutsiii, cd_distrito, ds_distrito, cd_concelho, ds_concelho, "
            "perc_valor_aprovado, perc_valor_pago) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    # --- Index PRR milestones ---
    print("\nIndexing PRR milestones...")
    for f in sorted(XLSX_DIR.glob("prr_milestones_*.xlsx")):
        print(f"  {f.name}: parsing...", end=" ", flush=True)
        rows = _parse_prr_milestones(f)
        if not rows:
            print("0 rows")
            continue
        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO prr_milestones "
            "(componente, data_ref, sequencial, codigo_reforma, designacao_reforma, "
            "tipo, designacao, indicador_qualitativo, indicador_quantitativo, "
            "referencia, objetivo, trimestre, ano, fonte_dados, responsabilidade, "
            "descricao, pressupostos_riscos, mecanismo_verificacao, "
            "indicadores_desembolso, natureza_medida, data_conclusao, valor_atingido) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    # --- Index budget execution ---
    print("\nIndexing budget execution...")
    for key in BUDGET_DATASETS:
        files = sorted(XLSX_DIR.glob(f"budget_{key}_*"))
        print(f"  [{key}] {len(files)} file(s) found")
        for f in files:
            print(f"    {f.name}: parsing...", end=" ", flush=True)
            rows = _parse_budget(f, key)
            if not rows:
                print("0 rows")
                continue
            changes_before = conn.total_changes
            conn.executemany(
                "INSERT OR IGNORE INTO budget "
                "(dataset_key, ano, mes, nivel_orcamental, descricao, "
                "valor_previsto, valor_realizado, percentagem) "
                "VALUES (?,?,?,?,?,?,?,?)",
                rows,
            )
            new_count = conn.total_changes - changes_before
            total_new += new_count
            print(f"{len(rows):,} rows ({new_count:,} new)")
            conn.commit()

    # Summary
    counts = {}
    for table in ("prr_contracts", "prr_entity_contracts", "prr_entities",
                   "prr_projects", "prr_locations", "prr_milestones", "budget"):
        counts[table] = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]

    print(f"\n  Index totals:")
    for table, count in counts.items():
        print(f"    {table}: {count:,}")
    print(f"    New rows this run: {total_new:,}")

    conn.close()


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def cmd_status(args):
    """Quick one-glance status overview."""
    if not DB_PATH.exists():
        print(f"  transparency.db: NOT FOUND")
        print(f"  Run 'python transparency_scraper.py download' then 'index'")
        return

    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    mtime = datetime.fromtimestamp(DB_PATH.stat().st_mtime)
    age_days = (datetime.now() - mtime).days

    conn = init_db()
    tables = ["prr_contracts", "prr_entities", "prr_projects", "prr_locations", "budget"]
    counts = {}
    for t in tables:
        try:
            counts[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        except Exception:
            counts[t] = 0
    conn.close()

    print(f"  transparency.db  {db_size:.1f} MB  ({age_days}d old)")
    for t, c in counts.items():
        label = t.replace("prr_", "prr ").replace("_", " ")
        print(f"    {label:<20} {c:>10,}")
    print()


def cmd_stats(args):
    """Show summary statistics."""
    conn = init_db()

    print("\n" + "=" * 70)
    print("  Transparency Data — Statistics")
    print("=" * 70)

    # PRR Contracts
    prr_total = conn.execute("SELECT COUNT(*) FROM prr_contracts").fetchone()[0]
    if prr_total:
        print(f"\n  PRR Contracts: {prr_total:,}")

        val = conn.execute(
            "SELECT SUM(COALESCE(montante, 0)) FROM prr_contracts"
        ).fetchone()
        print(f"    Total contract value: €{val[0]:,.0f}")

        # Top buyers (from entity-contract links)
        print(f"\n    Top Buyers (by contract count):")
        buyer_data = conn.execute(
            "SELECT ec.ds_entidade, ec.cd_entidade, COUNT(*), "
            "SUM(COALESCE(ec.valor_contrato, 0)) "
            "FROM prr_entity_contracts ec "
            "WHERE ec.ds_papel LIKE '%Comprador%' OR ec.ds_papel LIKE '%comprador%' "
            "OR ec.ds_papel LIKE '%Adjudicante%' "
            "GROUP BY ec.cd_entidade ORDER BY COUNT(*) DESC LIMIT 10"
        ).fetchall()
        if not buyer_data:
            # Fallback: all entities
            buyer_data = conn.execute(
                "SELECT ds_entidade, cd_entidade, COUNT(*), SUM(COALESCE(valor_contrato, 0)) "
                "FROM prr_entity_contracts WHERE ds_entidade != '' "
                "GROUP BY cd_entidade ORDER BY COUNT(*) DESC LIMIT 10"
            ).fetchall()
        for name, cd, cnt, val in buyer_data:
            print(f"      {name[:40]:40s} {cd:>10s}  {cnt:>4}  €{val:>12,.0f}")

        # Top suppliers
        print(f"\n    Top Suppliers (by contract value):")
        sup_data = conn.execute(
            "SELECT ec.ds_entidade, ec.cd_entidade, COUNT(*), "
            "SUM(COALESCE(ec.valor_contrato, 0)) "
            "FROM prr_entity_contracts ec "
            "WHERE ec.ds_papel LIKE '%Adjudicat%' "
            "GROUP BY ec.cd_entidade ORDER BY SUM(COALESCE(ec.valor_contrato, 0)) DESC LIMIT 10"
        ).fetchall()
        if not sup_data:
            sup_data = conn.execute(
                "SELECT ds_entidade, cd_entidade, COUNT(*), SUM(COALESCE(valor_contrato, 0)) "
                "FROM prr_entity_contracts WHERE ds_entidade != '' "
                "GROUP BY cd_entidade ORDER BY SUM(COALESCE(valor_contrato, 0)) DESC LIMIT 10"
            ).fetchall()
        for name, cd, cnt, val in sup_data:
            print(f"      {name[:40]:40s} {cd:>10s}  {cnt:>4}  €{val:>12,.0f}")

    else:
        print("\n  PRR Contracts: 0 (run 'index' first)")

    # PRR Entities
    ent_total = conn.execute("SELECT COUNT(*) FROM prr_entities").fetchone()[0]
    if ent_total:
        print(f"\n  PRR Entities: {ent_total:,}")

        # By role
        print(f"\n    By Role:")
        role_data = conn.execute(
            "SELECT papel, COUNT(*), SUM(COALESCE(valor_contratado, 0)) "
            "FROM prr_entities WHERE papel != '' "
            "GROUP BY papel ORDER BY COUNT(*) DESC"
        ).fetchall()
        for papel, cnt, val in role_data:
            print(f"      {papel[:40]:40s} {cnt:>5,}  €{val:>14,.0f}")

    # PRR Projects
    proj_total = conn.execute("SELECT COUNT(*) FROM prr_projects").fetchone()[0]
    if proj_total:
        print(f"\n  PRR Projects: {proj_total:,}")
        val = conn.execute(
            "SELECT SUM(COALESCE(valor_aprovado, 0)), SUM(COALESCE(valor_pago, 0)) "
            "FROM prr_projects"
        ).fetchone()
        print(f"    Total approved: €{val[0]:,.0f}")
        print(f"    Total paid: €{val[1]:,.0f}")
        if val[0] and val[0] > 0:
            print(f"    Execution rate: {val[1] / val[0] * 100:.1f}%")

    # PRR Locations
    loc_total = conn.execute("SELECT COUNT(*) FROM prr_locations").fetchone()[0]
    if loc_total:
        print(f"\n  PRR Locations: {loc_total:,}")

        # Top concelhos
        print(f"\n    Top Concelhos (by project count):")
        loc_data = conn.execute(
            "SELECT ds_concelho, COUNT(DISTINCT cd_projeto), "
            "SUM(COALESCE(perc_valor_aprovado, 0)) "
            "FROM prr_locations WHERE ds_concelho != '' "
            "GROUP BY ds_concelho ORDER BY COUNT(DISTINCT cd_projeto) DESC LIMIT 10"
        ).fetchall()
        for concelho, cnt, val in loc_data:
            print(f"      {concelho[:35]:35s} {cnt:>4} projects  €{val:>12,.0f}")

    # PRR Milestones
    ms_total = conn.execute("SELECT COUNT(*) FROM prr_milestones").fetchone()[0]
    if ms_total:
        print(f"\n  PRR Milestones: {ms_total:,}")

    # Budget
    budget_total = conn.execute("SELECT COUNT(*) FROM budget").fetchone()[0]
    if budget_total:
        print(f"\n  Budget Entries: {budget_total:,}")

        print(f"\n    By Year:")
        year_data = conn.execute(
            "SELECT ano, COUNT(*), SUM(COALESCE(valor_realizado, 0)) "
            "FROM budget WHERE ano IS NOT NULL "
            "GROUP BY ano ORDER BY ano"
        ).fetchall()
        for ano, cnt, val in year_data:
            print(f"      {ano}: {cnt:>6,} entries  €{val:>16,.0f}")

    print("\n" + "=" * 70)
    conn.close()


# ---------------------------------------------------------------------------
# PRR Analysis
# ---------------------------------------------------------------------------

def cmd_prr(args):
    """Analyze PRR contracts for corruption signals."""
    conn = init_db()

    prr_total = conn.execute("SELECT COUNT(*) FROM prr_contracts").fetchone()[0]
    if not prr_total:
        print("  PRR contracts index is empty. Run 'index' first.")
        conn.close()
        return

    print("\n" + "=" * 70)
    print("  PRR Contracts — Corruption Signal Analysis")
    print("=" * 70)

    # 1. High-value contracts
    print("\n  🔴 HIGH-VALUE PRR CONTRACTS (>€1M):")
    high_val = conn.execute(
        "SELECT c.cd_contrato, c.ds_contrato, c.montante, c.cd_projeto, "
        "c.dt_assinatura, c.cd_base_gov "
        "FROM prr_contracts c WHERE COALESCE(c.montante, 0) > 1000000 "
        "ORDER BY c.montante DESC LIMIT 20"
    ).fetchall()
    for cd, ds, montante, proj, dt, base in high_val:
        base_str = f"BASE#{base}" if base else "no BASE ref"
        print(f"    €{montante:>14,.0f}  [{cd}] {base_str}")
        print(f"      Project: {proj}")
        print(f"      Date: {dt}")
        if ds:
            print(f"      {ds[:70]}")
        print()

    # 2. Self-referencing entities
    print("\n  🔴 SELF-REFERENCING ENTITIES (same entity as buyer and supplier):")
    self_ref = conn.execute(
        "SELECT a.ds_entidade, a.cd_entidade, COUNT(*) as cnt, "
        "SUM(COALESCE(a.valor_contrato, 0)) as total_val "
        "FROM prr_entity_contracts a "
        "JOIN prr_entity_contracts b ON a.cd_contrato = b.cd_contrato "
        "WHERE a.cd_entidade = b.cd_entidade "
        "AND (a.ds_papel LIKE '%Comprador%' AND b.ds_papel LIKE '%Adjudicat%' "
        "     OR a.ds_papel LIKE '%Adjudicante%' AND b.ds_papel LIKE '%Adjudicat%') "
        "GROUP BY a.cd_entidade HAVING cnt >= 1 "
        "ORDER BY total_val DESC LIMIT 10"
    ).fetchall()
    if self_ref:
        for name, cd, cnt, val in self_ref:
            print(f"    {name[:45]:45s} {cnt} contracts  €{val:,.0f}")
    else:
        print("    None found (or data structure doesn't support this analysis)")

    # 3. Concentrated suppliers
    print("\n  🟠 SUPPLIER CONCENTRATION (same supplier, 3+ PRR contracts):")
    supplier_conc = conn.execute(
        "SELECT ec.ds_entidade, ec.cd_entidade, COUNT(*) as cnt, "
        "SUM(COALESCE(ec.valor_contrato, 0)) as total_val "
        "FROM prr_entity_contracts ec "
        "WHERE ec.ds_papel LIKE '%Adjudicat%' "
        "GROUP BY ec.cd_entidade HAVING cnt >= 3 "
        "ORDER BY total_val DESC LIMIT 15"
    ).fetchall()
    if not supplier_conc:
        supplier_conc = conn.execute(
            "SELECT ds_entidade, cd_entidade, COUNT(*) as cnt, "
            "SUM(COALESCE(valor_contrato, 0)) as total_val "
            "FROM prr_entity_contracts "
            "GROUP BY cd_entidade HAVING cnt >= 3 "
            "ORDER BY total_val DESC LIMIT 15"
        ).fetchall()
    for name, cd, cnt, val in supplier_conc:
        print(f"    {name[:40]:40s} {cd:>10s}")
        print(f"      {cnt} contracts, €{val:,.0f}")

    # 4. Projects with low execution rate
    print("\n  🟡 PROJECTS WITH LOW EXECUTION (<50% paid):")
    low_exec = conn.execute(
        "SELECT cd_projeto, ds_projeto, valor_aprovado, valor_pago, "
        "CASE WHEN valor_aprovado > 0 THEN valor_pago * 100.0 / valor_aprovado ELSE 0 END as pct "
        "FROM prr_projects "
        "WHERE valor_aprovado > 1000000 AND valor_pago IS NOT NULL "
        "AND CASE WHEN valor_aprovado > 0 THEN valor_pago * 100.0 / valor_aprovado ELSE 0 END < 50 "
        "ORDER BY valor_aprovado DESC LIMIT 10"
    ).fetchall()
    for pid, titulo, aprov, pago, pct in low_exec:
        print(f"    {pid}: {pct:.1f}% executed (€{pago:,.0f} / €{aprov:,.0f})")
        if titulo:
            print(f"      {titulo[:60]}")

    # 5. Geographic concentration
    print("\n  🟡 GEOGRAPHIC CONCENTRATION:")
    geo = conn.execute(
        "SELECT ds_concelho, COUNT(DISTINCT cd_projeto) as projects, "
        "SUM(COALESCE(perc_valor_aprovado, 0)) as total_val "
        "FROM prr_locations WHERE ds_concelho != '' "
        "GROUP BY ds_concelho ORDER BY total_val DESC LIMIT 10"
    ).fetchall()
    for concelho, projects, val in geo:
        print(f"    {concelho[:35]:35s} {projects:>4} projects  €{val:>12,.0f}")

    print("\n" + "=" * 70)
    conn.close()


# ---------------------------------------------------------------------------
# Budget Analysis
# ---------------------------------------------------------------------------

def cmd_budget(args):
    """Analyze budget execution for anomalies."""
    conn = init_db()

    budget_total = conn.execute("SELECT COUNT(*) FROM budget").fetchone()[0]
    if not budget_total:
        print("  Budget index is empty. Run 'index' first.")
        conn.close()
        return

    print("\n" + "=" * 70)
    print("  Budget Execution — Analysis")
    print("=" * 70)

    # Budget variance
    print("\n  🔴 SIGNIFICANT BUDGET VARIANCES (>50% over/under):")
    variance = conn.execute(
        "SELECT descricao, ano, valor_previsto, valor_realizado, "
        "CASE WHEN valor_previsto > 0 "
        "     THEN (valor_realizado - valor_previsto) * 100.0 / valor_previsto "
        "     ELSE 0 END as pct_diff "
        "FROM budget "
        "WHERE valor_previsto > 1000000 AND valor_realizado > 0 "
        "AND ABS(CASE WHEN valor_previsto > 0 "
        "     THEN (valor_realizado - valor_previsto) * 100.0 / valor_previsto "
        "     ELSE 0 END) > 50 "
        "ORDER BY ABS(pct_diff) DESC LIMIT 20"
    ).fetchall()
    for desc, ano, prev, real, pct in variance:
        direction = "OVER" if pct > 0 else "UNDER"
        print(f"    {direction} {abs(pct):>6.1f}%  €{real:>14,.0f} vs €{prev:>14,.0f}")
        print(f"      [{ano}] {(desc or '')[:60]}")

    # Year-over-year trends
    print("\n  📈 SPENDING TRENDS BY YEAR:")
    year_trend = conn.execute(
        "SELECT ano, SUM(COALESCE(valor_previsto, 0)) as prev, "
        "SUM(COALESCE(valor_realizado, 0)) as real "
        "FROM budget WHERE ano IS NOT NULL AND valor_realizado > 0 "
        "GROUP BY ano ORDER BY ano"
    ).fetchall()
    prev_real = None
    for ano, prev, real in year_trend:
        yoy = ""
        if prev_real and prev_real > 0:
            pct = (real - prev_real) / prev_real * 100
            yoy = f"({pct:+.1f}%)"
        prev_real = real
        exec_pct = (real / prev * 100) if prev and prev > 0 else 0
        print(f"    {ano}: €{real:>16,.0f} (exec: {exec_pct:.1f}%) {yoy}")

    print("\n" + "=" * 70)
    conn.close()


# ---------------------------------------------------------------------------
# Cross-reference
# ---------------------------------------------------------------------------

def cmd_crossref(args):
    """Cross-reference PRR entities with procurement.db.

    Uses a single aggregated query instead of N+1 for performance.
    Matches PRR entities as buyers (adjudicante_nif) and as suppliers
    (NIF appearing in adjudicatarios text).
    """
    conn = init_db()

    if not PROCUREMENT_DB.exists():
        print(f"  procurement.db not found at {PROCUREMENT_DB}")
        conn.close()
        return

    prr_total = conn.execute("SELECT COUNT(*) FROM prr_contracts").fetchone()[0]
    if not prr_total:
        print("  PRR contracts index is empty. Run 'index' first.")
        conn.close()
        return

    # Open procurement.db (PRAGMA-tuned via utils_db.connect).
    proc_conn = db_connect(str(PROCUREMENT_DB))
    tables = [r[0] for r in proc_conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]

    if "contratos" not in tables:
        print(f"  No 'contratos' table in procurement.db (available: {tables})")
        proc_conn.close()
        conn.close()
        return

    print("\n" + "=" * 70)
    print("  PRR × Procurement Cross-Reference")
    print("=" * 70)

    # --- Match 1: PRR entities as BUYERS (single aggregated query) ---
    # procurement.db stores buyer NIF in adjudicante_nif
    buyer_rows = proc_conn.execute(
        "SELECT adjudicante_nif, COUNT(*) as cnt "
        "FROM contratos WHERE adjudicante_nif != '-' AND adjudicante_nif != '' "
        "GROUP BY adjudicante_nif"
    ).fetchall()
    buyer_nif_counts = {nif: cnt for nif, cnt in buyer_rows}
    print(f"\n  Unique buyer NIFs in procurement.db: {len(buyer_nif_counts):,}")

    # Match against PRR entities
    prr_entities = conn.execute(
        "SELECT cd_entidade, ds_entidade, nif, valor_contratado, papel "
        "FROM prr_entities WHERE nif != ''"
    ).fetchall()
    print(f"  PRR entities with NIF: {len(prr_entities):,}")

    buyer_matches = []
    for cd_ent, name, nif, prr_val, papel in prr_entities:
        if nif in buyer_nif_counts:
            buyer_matches.append({
                "cd_ent": cd_ent, "name": name, "nif": nif,
                "prr_value": prr_val or 0, "proc_count": buyer_nif_counts[nif],
                "papel": papel,
            })

    buyer_matches.sort(key=lambda x: -(x["prr_value"]))
    print(f"\n  PRR entities as buyers in BASE.gov.pt: {len(buyer_matches):,}")
    for m in buyer_matches[:15]:
        print(f"    {m['name'][:45]:45s} NIF={m['nif']}  PRR €{m['prr_value']:>12,.0f}  BASE {m['proc_count']:>5} contracts  [{m['papel']}]")

    # --- Match 2: PRR entity NIFs in adjudicatarios text ---
    # Use the indexed ``adjudicatario_nif`` column (populated by
    # ``add_adjudicatario_nif.py``). The column stores the FIRST 9-digit
    # NIF from each contract's ``adjudicatarios`` text. Multi-NIF joint
    # ventures are not fully captured — add a ``contrato_nif`` junction
    # table if joint-venture coverage is required (see ROADMAP P3.3).
    print(f"\n  --- Match 2: PRR entity NIFs in adjudicatarios text ---")
    supplier_nif_counts = {}
    t_nif = time.time()
    rows = proc_conn.execute("""
        SELECT adjudicatario_nif, COUNT(*) AS cnt
        FROM contratos
        WHERE adjudicatario_nif IS NOT NULL AND adjudicatario_nif != ''
        GROUP BY adjudicatario_nif
    """).fetchall()
    supplier_nif_counts = {r["adjudicatario_nif"]: r["cnt"] for r in rows}
    print(f"  Unique supplier NIFs (indexed first-NIF lookup): {len(supplier_nif_counts):,} in {time.time() - t_nif:.2f}s")

    supplier_matches = []
    for cd_ent, name, nif, prr_val, papel in prr_entities:
        if nif in supplier_nif_counts:
            supplier_matches.append({
                "cd_ent": cd_ent, "name": name, "nif": nif,
                "prr_value": prr_val or 0, "proc_count": supplier_nif_counts[nif],
                "papel": papel,
            })

    supplier_matches.sort(key=lambda x: -(x["prr_value"]))
    print(f"  PRR entities as suppliers in BASE.gov.pt: {len(supplier_matches):,}")
    for m in supplier_matches[:15]:
        print(f"    {m['name'][:45]:45s} NIF={m['nif']}  PRR €{m['prr_value']:>12,.0f}  BASE {m['proc_count']:>5} contracts  [{m['papel']}]")

    # --- Summary ---
    all_nifs = set(m["nif"] for m in buyer_matches) | set(m["nif"] for m in supplier_matches)
    print(f"\n  --- Summary ---")
    print(f"  Total unique PRR entities found in procurement.db: {len(all_nifs):,}")
    print(f"  As buyers: {len(buyer_matches):,}")
    print(f"  As suppliers: {len(supplier_matches):,}")

    print("\n" + "=" * 70)
    proc_conn.close()
    conn.close()


# ---------------------------------------------------------------------------
# Entity search
# ---------------------------------------------------------------------------

def cmd_entity(args):
    """Show PRR data for a specific entity."""
    conn = init_db()

    nif = args.nif
    if not nif:
        print("  Specify --nif")
        conn.close()
        return

    # Find entity info
    entity = conn.execute(
        "SELECT ds_entidade, papel, localizacao, valor_contratado, valor_pago "
        "FROM prr_entities WHERE nif = ? LIMIT 1",
        (nif,),
    ).fetchone()
    entity_name = entity[0] if entity else nif

    print("\n" + "=" * 70)
    print(f"  Entity: {entity_name} (NIF: {nif})")
    if entity:
        print(f"  Role: {entity[1]} | Location: {entity[2]}")
        print(f"  Contracted: €{entity[3]:,.0f} | Paid: €{entity[4]:,.0f}")
    print("=" * 70)

    # Contracts involving this entity
    contracts = conn.execute(
        "SELECT ec.cd_contrato, ec.ds_contrato, ec.ds_papel, ec.valor_contrato, "
        "c.cd_projeto, c.ds_projeto, c.dt_assinatura "
        "FROM prr_entity_contracts ec "
        "LEFT JOIN prr_contracts c ON ec.cd_contrato = c.cd_contrato "
        "WHERE ec.cd_entidade IN ("
        "  SELECT cd_entidade FROM prr_entities WHERE nif = ?"
        ") ORDER BY COALESCE(ec.valor_contrato, 0) DESC LIMIT ?",
        (nif, args.limit),
    ).fetchall()

    if contracts:
        print(f"\n  Contracts ({len(contracts):,}):")
        for cd, ds, papel, val, proj, proj_ds, dt in contracts:
            val_str = f"€{val:,.0f}" if val else "N/A"
            print(f"    [{cd}] {val_str} — {papel}")
            print(f"      Project: {proj} — {(proj_ds or '')[:50]}")
            print(f"      Date: {dt}")
            if ds:
                print(f"      {ds[:60]}")

    # Projects
    cd_ent = entity[0] if entity else None
    if cd_ent:
        projects = conn.execute(
            "SELECT cd_projeto, ds_projeto, valor_aprovado, valor_pago "
            "FROM prr_projects WHERE cd_projeto IN ("
            "  SELECT cd_projeto FROM prr_entities WHERE nif = ?"
            ") ORDER BY valor_aprovado DESC LIMIT ?",
            (nif, args.limit),
        ).fetchall()
        if projects:
            print(f"\n  Projects ({len(projects):,}):")
            for pid, titulo, aprov, pago in projects:
                aprov_str = f"€{aprov:,.0f}" if aprov else "N/A"
                pct = (pago / aprov * 100) if aprov and pago else 0
                print(f"    [{pid}] {aprov_str} ({pct:.1f}% paid)")
                if titulo:
                    print(f"      {titulo[:60]}")

    print("\n" + "=" * 70)
    conn.close()


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

def cmd_search(args):
    """Full-text search in PRR data."""
    conn = init_db()

    query = f"%{args.query}%"

    # Search contracts
    rows = conn.execute(
        "SELECT cd_contrato, ds_contrato, montante, cd_projeto, dt_assinatura "
        "FROM prr_contracts WHERE ds_contrato LIKE ? OR sumario LIKE ? "
        "ORDER BY montante DESC LIMIT ?",
        (query, query, args.limit),
    ).fetchall()

    if rows:
        print(f"\n  PRR Contracts matching '{args.query}':\n")
        for cd, ds, montante, proj, dt in rows:
            val_str = f"€{montante:,.0f}" if montante else "N/A"
            print(f"  [{cd}] {val_str} — Project: {proj}")
            if ds:
                print(f"    {ds[:70]}")
            print()

    # Search entities
    ent_rows = conn.execute(
        "SELECT cd_entidade, ds_entidade, nif, papel, valor_contratado "
        "FROM prr_entities WHERE ds_entidade LIKE ? OR atividade_economica LIKE ? "
        "ORDER BY valor_contratado DESC LIMIT ?",
        (query, query, args.limit),
    ).fetchall()

    if ent_rows:
        print(f"\n  PRR Entities matching '{args.query}':\n")
        for cd, name, nif, papel, val in ent_rows:
            val_str = f"€{val:,.0f}" if val else "N/A"
            print(f"  [{cd}] {name[:45]} NIF={nif} — {papel}")
            print(f"    Contracted: {val_str}")

    # Search projects
    proj_rows = conn.execute(
        "SELECT cd_projeto, ds_projeto, valor_aprovado, valor_pago "
        "FROM prr_projects WHERE ds_projeto LIKE ? OR sumario LIKE ? "
        "ORDER BY valor_aprovado DESC LIMIT ?",
        (query, query, args.limit),
    ).fetchall()

    if proj_rows:
        print(f"\n  PRR Projects matching '{args.query}':\n")
        for pid, titulo, aprov, pago in proj_rows:
            aprov_str = f"€{aprov:,.0f}" if aprov else "N/A"
            print(f"  [{pid}] {aprov_str}")
            if titulo:
                print(f"    {titulo[:70]}")
            print()

    if not rows and not ent_rows and not proj_rows:
        print("  No results found.")

    conn.close()


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def cmd_export(args):
    """Export index to JSON."""
    conn = init_db()

    data = {}

    # Export PRR contracts
    rows = conn.execute(
        "SELECT cd_contrato, ds_contrato, sumario, cd_base_gov, dt_assinatura, "
        "montante, cd_projeto, ds_projeto "
        "FROM prr_contracts ORDER BY montante DESC"
    ).fetchall()
    data["prr_contracts"] = [
        {"cd_contrato": r[0], "ds_contrato": r[1], "sumario": r[2],
         "cd_base_gov": r[3], "dt_assinatura": r[4], "montante": r[5],
         "cd_projeto": r[6], "ds_projeto": r[7]}
        for r in rows
    ]

    # Export PRR entities
    rows = conn.execute(
        "SELECT cd_entidade, nif, ds_entidade, papel, atividade_economica, "
        "localizacao, valor_contratado, valor_pago "
        "FROM prr_entities ORDER BY valor_contratado DESC"
    ).fetchall()
    data["prr_entities"] = [
        {"cd_entidade": r[0], "nif": r[1], "ds_entidade": r[2], "papel": r[3],
         "atividade_economica": r[4], "localizacao": r[5],
         "valor_contratado": r[6], "valor_pago": r[7]}
        for r in rows
    ]

    # Export PRR projects
    rows = conn.execute(
        "SELECT cd_projeto, ds_projeto, sumario, valor_aprovado, valor_pago, "
        "dt_inicio, dt_prevista_conclusao "
        "FROM prr_projects ORDER BY valor_aprovado DESC"
    ).fetchall()
    data["prr_projects"] = [
        {"cd_projeto": r[0], "ds_projeto": r[1], "sumario": r[2],
         "valor_aprovado": r[3], "valor_pago": r[4],
         "dt_inicio": r[5], "dt_prevista_conclusao": r[6]}
        for r in rows
    ]

    # Export PRR locations
    rows = conn.execute(
        "SELECT cd_projeto, ds_concelho, ds_distrito, perc_valor_aprovado, perc_valor_pago "
        "FROM prr_locations ORDER BY perc_valor_aprovado DESC"
    ).fetchall()
    data["prr_locations"] = [
        {"cd_projeto": r[0], "ds_concelho": r[1], "ds_distrito": r[2],
         "perc_valor_aprovado": r[3], "perc_valor_pago": r[4]}
        for r in rows
    ]

    out_path = Path(args.out) if args.out else DATA_DIR / "transparency_export.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"  Exported:")
    for key, items in data.items():
        print(f"    {key}: {len(items):,}")
    print(f"  To: {out_path}")

    conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Transparency Scraper — Portuguese budget execution and PRR data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    sub = parser.add_subparsers(dest="command")

    # Download
    dl = sub.add_parser("download", help="Download datasets from dados.gov.pt")
    dl.add_argument("--type", choices=["all", "budget", "prr"], default="all",
                    help="Type of data to download (default: all)")

    # Index
    idx = sub.add_parser("index", help="Parse XLSX files into SQLite")
    idx.add_argument("--force", action="store_true", help="Re-index from scratch")

    # Status
    sub.add_parser("status", help="Quick one-glance status overview")

    # Stats
    sub.add_parser("stats", help="Summary statistics")

    # PRR analysis
    sub.add_parser("prr", help="PRR contract corruption signal analysis")

    # Budget analysis
    sub.add_parser("budget", help="Budget execution analysis")

    # Cross-reference
    sub.add_parser("crossref", help="Cross-ref with procurement.db")

    # Entity
    ent = sub.add_parser("entity", help="Entity-specific analysis")
    ent.add_argument("--nif", required=True, help="Entity NIF")
    ent.add_argument("--limit", type=int, default=20)

    # Search
    srch = sub.add_parser("search", help="Full-text search PRR data")
    srch.add_argument("--query", required=True, help="Search query")
    srch.add_argument("--limit", type=int, default=20)

    # Export
    exp = sub.add_parser("export", help="Export to JSON")
    exp.add_argument("--out", help="Output file path")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "download": cmd_download,
        "index": cmd_index,
        "status": cmd_status,
        "stats": cmd_stats,
        "prr": cmd_prr,
        "budget": cmd_budget,
        "crossref": cmd_crossref,
        "entity": cmd_entity,
        "search": cmd_search,
        "export": cmd_export,
    }

    commands[args.command](args)


if __name__ == "__main__":
    main()
