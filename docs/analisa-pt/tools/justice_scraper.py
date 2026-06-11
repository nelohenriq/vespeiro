#!/usr/bin/env python3
"""Justice Data Scraper — Download and analyze Portuguese court statistics.

Fetches justice datasets from dados.justica.gov.pt (CKAN portal), stores
in SQLite, and provides analysis commands for corruption detection.

Covers: court case movements, corruption/money-laundering prosecutions,
case duration, prison population, and nationality processes.

Usage:
    python justice_scraper.py list               # List all available datasets
    python justice_scraper.py download           # Download target datasets
    python justice_scraper.py download --all     # Download all 62 datasets
    python justice_scraper.py index              # Parse CSV into SQLite
    python justice_scraper.py status             # Quick status
    python justice_scraper.py stats              # Summary statistics
    python justice_scraper.py corruption         # Corruption case analysis
    python justice_scraper.py courts             # Court case flow analysis
    python justice_scraper.py prisons            # Prison population analysis
    python justice_scraper.py query "SELECT ..." # Run SQL
    python justice_scraper.py export --out X     # Export JSON
"""

import sys
import json
import csv
import io
import sqlite3
import argparse
import time
from pathlib import Path
from datetime import datetime, timezone

try:
    import urllib.request
    import ssl
except ImportError:
    print("ERROR: urllib required (built-in)")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
DB_PATH = DATA_DIR / "justice.db"
RAW_DIR = DATA_DIR / "justice_raw"

# SSL context
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json,text/csv,*/*",
}

# CKAN API base
CKAN_API = "https://dados.justica.gov.pt/api/3/action"

# Números da Justiça — supplementary direct-download files
# These are static files hosted on numeros.justica.gov.pt that complement
# the CKAN API data. URLs may change; failures are non-fatal.
SUPPLEMENTARY_SOURCES = {
    "nj_sintese_transacao": {
        "url": "https://numeros.justica.gov.pt/sites/default/files/2025-12/sintese-transacao_dgpj.xlsx",
        "title": "Síntese Transação (DGPJ)",
        "description": "Transaction/plea bargaining synthesis from DGPJ",
        "category": "courts",
    },
}



# Target datasets for corruption/justice analysis
TARGET_DATASETS = {
    "corrupcaopj": {
        "title": "Processos de crime de corrupção",
        "description": "Criminal processes for corruption investigated by PJ",
        "category": "corruption",
    },
    "branqueamentopj": {
        "title": "Processos de branqueamento de capitais",
        "description": "Money laundering criminal processes investigated by PJ",
        "category": "corruption",
    },
    "movimtribjud1instancia": {
        "title": "Movimento Tribunais Judiciais 1ª Instância",
        "description": "Cases filed, finalized, and pending in 1st instance courts",
        "category": "courts",
    },
    "duracaomedia": {
        "title": "Duração Média Processos Findos",
        "description": "Average duration in months of finalized cases in 1st instance courts",
        "category": "courts",
    },
    "processossaidosarquivtribunais": {
        "title": "Processos nos arquivos dos Tribunais",
        "description": "Cases entering and leaving court archives",
        "category": "courts",
    },
    "procentraprocsai": {
        "title": "Processos entrados/saídos arquivos",
        "description": "Cases entering vs leaving court archives",
        "category": "courts",
    },
    "reclusos": {
        "title": "Reclusos",
        "description": "Prison population by sex (annual)",
        "category": "prisons",
    },
    "nacionalidade": {
        "title": "Processos de nacionalidade",
        "description": "Nationality processes concluded by month",
        "category": "nationality",
    },
    "registocriminal": {
        "title": "Registo Criminal Online",
        "description": "Online criminal record requests paid",
        "category": "criminal_records",
    },
    "medidas": {
        "title": "Medidas Simplex Justiça",
        "description": "Justice Simplex measures 2016-2017",
        "category": "justice_measures",
    },
}


# ---------------------------------------------------------------------------
# CKAN API Client
# ---------------------------------------------------------------------------

def ckan_api(endpoint: str, params: dict = None) -> dict:
    """Call the CKAN API."""
    url = f"{CKAN_API}/{endpoint}"
    if params:
        query = "&".join(f"{k}={v}" for k, v in params.items())
        url += f"?{query}"

    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=30, context=SSL_CTX)
    data = json.loads(resp.read())
    if not data.get("success"):
        raise RuntimeError(f"CKAN API error: {data}")
    return data.get("result", {})


def list_datasets(search: str = None) -> list[dict]:
    """List all available datasets."""
    if search:
        result = ckan_api("package_search", {"q": search, "rows": "50"})
        return result.get("results", [])
    else:
        names = ckan_api("package_list")
        datasets = []
        for name in names:
            try:
                pkg = ckan_api("package_show", {"id": name})
                datasets.append(pkg)
            except Exception:
                pass
            time.sleep(0.1)
        return datasets


def get_dataset_resources(dataset_name: str) -> list[dict]:
    """Get download resources for a dataset, preferring CSV."""
    pkg = ckan_api("package_show", {"id": dataset_name})
    resources = pkg.get("resources", [])

    # Prefer CSV, then XLSX, then JSON
    csv_resources = [r for r in resources if r.get("format", "").upper() == "CSV"]
    if csv_resources:
        return csv_resources

    xlsx_resources = [r for r in resources if r.get("format", "").upper() in ("XLSX", "XLS")]
    if xlsx_resources:
        return xlsx_resources

    return resources


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def init_db(force: bool = False) -> sqlite3.Connection:
    """Initialize the justice database."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    if force and DB_PATH.exists():
        DB_PATH.unlink()
        print("  Deleted existing database.")

    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")

    # Generic data table for all datasets
    conn.execute("""
        CREATE TABLE IF NOT EXISTS justice_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset TEXT NOT NULL,
            category TEXT,
            row_data TEXT,
            UNIQUE(dataset, row_data)
        )
    """)

    # Dataset metadata
    conn.execute("""
        CREATE TABLE IF NOT EXISTS datasets (
            name TEXT PRIMARY KEY,
            title TEXT,
            description TEXT,
            category TEXT,
            resource_url TEXT,
            resource_format TEXT,
            row_count INTEGER DEFAULT 0,
            downloaded_at TEXT,
            indexed_at TEXT
        )
    """)

    # Corruption-specific table (parsed from corrupcaopj)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS corruption_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset TEXT,
            year INTEGER,
            period TEXT,
            value REAL,
            category TEXT,
            dimension TEXT,
            raw_row TEXT,
            UNIQUE(dataset, year, period, dimension)
        )
    """)

    # Court movements table (parsed from movimtribjud1instancia)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS court_movements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            court_type TEXT,
            period TEXT,
            year INTEGER,
            entered INTEGER,
            finalized INTEGER,
            pending INTEGER,
            raw_row TEXT
        )
    """)

    # Prison population table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prison_population (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER,
            sex TEXT,
            count INTEGER,
            raw_row TEXT,
            UNIQUE(year, sex)
        )
    """)

    # Indexes
    conn.execute("CREATE INDEX IF NOT EXISTS idx_jd_dataset ON justice_data(dataset)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cc_year ON corruption_cases(year)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cm_year ON court_movements(year)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_pp_year ON prison_population(year)")

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def _download_file(url: str, local_path: Path, label: str = "") -> int:
    """Download a file from URL to local path.

    Returns the number of bytes downloaded (0 if already exists or failed).
    """
    if local_path.exists() and local_path.stat().st_size > 100:
        size = local_path.stat().st_size
        print(f"    {label}: already exists ({size:,} bytes)")
        return 0

    print(f"    {label}: downloading...", end=" ", flush=True)
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=60, context=SSL_CTX)
        data = resp.read()
        local_path.write_bytes(data)
        print(f"done ({len(data):,} bytes)")
        return len(data)
    except Exception as e:
        print(f"FAILED: {e}")
        return 0


def cmd_download(args):
    """Download justice datasets from dados.justica.gov.pt."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    download_all = getattr(args, "all", False)
    target_name = getattr(args, "dataset", None)

    if download_all:
        print("\n  Listing all datasets from CKAN API...")
        try:
            names = ckan_api("package_list")
            print(f"  Found {len(names)} datasets")
        except Exception as e:
            print(f"  ERROR listing datasets: {e}")
            return
    elif target_name:
        names = [target_name]
    else:
        names = list(TARGET_DATASETS.keys())
        print(f"\n  Downloading {len(names)} target justice datasets...")

    total_downloaded = 0
    total_size = 0

    for name in names:
        info = TARGET_DATASETS.get(name, {"title": name, "category": ""})
        print(f"\n  [{name}] {info.get('title', name)}")

        try:
            resources = get_dataset_resources(name)
        except Exception as e:
            print(f"    ERROR: {e}")
            continue

        if not resources:
            print(f"    No downloadable resources found")
            continue

        for r in resources:
            fmt = r.get("format", "?").upper()
            url = r.get("url", "")
            rname = r.get("name", r.get("description", fmt))

            if not url:
                continue

            local_path = RAW_DIR / f"{name}.{fmt.lower()}"
            size = _download_file(url, local_path, fmt)
            if size > 0:
                total_downloaded += 1
                total_size += size
            time.sleep(0.3)

    # --- Supplementary sources (numeros.justica.gov.pt) ---
    print(f"\n  Supplementary sources (numeros.justica.gov.pt)...")
    for name, info in SUPPLEMENTARY_SOURCES.items():
        url = info["url"]
        fmt = url.rsplit(".", 1)[-1].upper()
        local_path = RAW_DIR / f"{name}.{fmt.lower()}"
        size = _download_file(url, local_path, name)
        if size > 0:
            total_downloaded += 1
            total_size += size
        time.sleep(0.3)

    print(f"\n  Total: {total_downloaded} files, {total_size:,} bytes")
    if total_downloaded > 0:
        print(f"  Run 'python justice_scraper.py index' to build the database")
    print()


# ---------------------------------------------------------------------------
# Index
# ---------------------------------------------------------------------------

def cmd_index(args):
    """Parse downloaded CSV/XLSX files into SQLite."""
    conn = init_db(force=getattr(args, "force", False))

    if getattr(args, "force", False):
        conn.execute("DELETE FROM justice_data")
        conn.execute("DELETE FROM corruption_cases")
        conn.execute("DELETE FROM court_movements")
        conn.execute("DELETE FROM prison_population")
        conn.execute("DELETE FROM datasets")
        conn.commit()
        print("  Cleared existing data.\n")

    total_rows = 0

    for f in sorted(RAW_DIR.glob("*")):
        if not f.is_file():
            continue

        name = f.stem
        fmt = f.suffix.lstrip(".").upper()
        # Check both TARGET_DATASETS and SUPPLEMENTARY_SOURCES
        info = TARGET_DATASETS.get(name, SUPPLEMENTARY_SOURCES.get(name, {"title": name, "category": ""}))
        category = info.get("category", "")

        print(f"  {f.name}: parsing...", end=" ", flush=True)

        try:
            if fmt == "CSV":
                rows = _parse_csv(f)
            elif fmt in ("XLSX", "XLS"):
                rows = _parse_xlsx(f)
            else:
                print(f"skipped (unsupported format: {fmt})")
                continue
        except Exception as e:
            print(f"FAILED: {e}")
            continue

        if not rows:
            print("0 rows")
            continue

        # Store raw rows
        changes_before = conn.total_changes
        for row in rows:
            row_json = json.dumps(row, ensure_ascii=False, default=str)
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO justice_data (dataset, category, row_data) VALUES (?, ?, ?)",
                    (name, category, row_json),
                )
            except sqlite3.Error:
                pass

        new_count = conn.total_changes - changes_before
        total_rows += new_count
        print(f"{len(rows):,} rows ({new_count:,} new)")

        # Parse into typed tables for specific datasets
        if name == "corrupcaopj" or name == "branqueamentopj":
            _index_corruption(conn, name, category, rows)
        elif name == "movimtribjud1instancia":
            _index_court_movements(conn, rows)
        elif name == "reclusos":
            _index_prison(conn, rows)

        conn.commit()

        # Update dataset metadata
        conn.execute("""
            INSERT OR REPLACE INTO datasets
            (name, title, description, category, row_count, indexed_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            name, info.get("title", name), info.get("description", ""),
            category, len(rows), datetime.now(timezone.utc).isoformat(),
        ))
        conn.commit()

    count = conn.execute("SELECT COUNT(*) FROM justice_data").fetchone()[0]
    n_ds = conn.execute("SELECT COUNT(DISTINCT dataset) FROM justice_data").fetchone()[0]

    print(f"\n  Index totals:")
    print(f"    observations:  {count:,}")
    print(f"    datasets:      {n_ds}")
    print(f"    New rows:      {total_rows:,}")

    conn.close()


def _parse_csv(path: Path) -> list[dict]:
    """Parse a CSV file into list of dicts.

    Tries utf-8 first (strict), then falls back to latin-1/cp1252 if decoding fails.
    """
    raw = path.read_bytes()
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            content = raw.decode(encoding)
            reader = csv.DictReader(io.StringIO(content))
            return [dict(row) for row in reader]
        except (UnicodeDecodeError, UnicodeError):
            continue
    # Last resort: replace undecodable bytes
    content = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(content))
    return [dict(row) for row in reader]


def _parse_xlsx(path: Path) -> list[dict]:
    """Parse an XLSX file into list of dicts."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed", end=" ")
        return []

    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    headers = [str(cell.value or f"col_{i}") for i, cell in enumerate(ws[1])]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(v is not None for v in row):
            rows.append({h: v for h, v in zip(headers, row)})
    wb.close()
    return rows


def _index_corruption(conn, dataset, category, rows):
    """Parse corruption/money-laundering data into corruption_cases table."""
    for row in rows:
        # Try to extract year and value from common column patterns
        year = None
        value = None
        dimension = ""
        period = ""

        for k, v in row.items():
            kl = k.lower().strip() if k else ""
            if "ano" in kl or "year" in kl:
                try:
                    year = int(v)
                except (ValueError, TypeError):
                    pass
            elif "valor" in kl or "total" in kl or "nº" in kl or "processos" in kl:
                try:
                    value = float(v)
                except (ValueError, TypeError):
                    pass
            elif "mês" in kl or "mes" in kl or "periodo" in kl or "trimestre" in kl:
                period = str(v or "")
            elif "tipo" in kl or "crime" in kl or "categoria" in kl:
                dimension = str(v or "")

        if year or value:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO corruption_cases
                    (dataset, year, period, value, category, dimension, raw_row)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    dataset, year, period, value, category, dimension,
                    json.dumps(row, ensure_ascii=False, default=str),
                ))
            except sqlite3.Error:
                pass


def _index_court_movements(conn, rows):
    """Parse court movement data into court_movements table."""
    for row in rows:
        court_type = ""
        period = ""
        year = None
        entered = None
        finalized = None
        pending = None

        for k, v in row.items():
            kl = k.lower().strip() if k else ""
            if "tribunal" in kl or "jurisdiç" in kl or "tipo" in kl:
                court_type = str(v or "")
            elif "ano" in kl or "year" in kl:
                try:
                    year = int(v)
                except (ValueError, TypeError):
                    pass
            elif "entrado" in kl or "entrada" in kl:
                try:
                    entered = int(v)
                except (ValueError, TypeError):
                    pass
            elif "findo" in kl or "fini" in kl or "saído" in kl:
                try:
                    finalized = int(v)
                except (ValueError, TypeError):
                    pass
            elif "pendente" in kl:
                try:
                    pending = int(v)
                except (ValueError, TypeError):
                    pass
            elif "mês" in kl or "periodo" in kl:
                period = str(v or "")

        try:
            conn.execute("""
                INSERT INTO court_movements
                (court_type, period, year, entered, finalized, pending, raw_row)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                court_type, period, year, entered, finalized, pending,
                json.dumps(row, ensure_ascii=False, default=str),
            ))
        except sqlite3.Error:
            pass


def _index_prison(conn, rows):
    """Parse prison population data into prison_population table."""
    for row in rows:
        year = None
        sex = ""
        count = None

        for k, v in row.items():
            kl = k.lower().strip() if k else ""
            if "ano" in kl or "year" in kl:
                try:
                    year = int(v)
                except (ValueError, TypeError):
                    pass
            elif "sexo" in kl or "sex" in kl or "género" in kl:
                sex = str(v or "")
            elif "total" in kl or "reclusos" in kl or "nº" in kl:
                try:
                    count = int(v)
                except (ValueError, TypeError):
                    pass

        if year is not None:
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO prison_population
                    (year, sex, count, raw_row)
                    VALUES (?, ?, ?, ?)
                """, (
                    year, sex, count,
                    json.dumps(row, ensure_ascii=False, default=str),
                ))
            except sqlite3.Error:
                pass


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

def cmd_list(args):
    """List available datasets."""
    print(f"\n  Fetching dataset list from dados.justica.gov.pt...")
    try:
        names = ckan_api("package_list")
    except Exception as e:
        print(f"  ERROR: {e}")
        return

    print(f"  Found {len(names)} datasets\n")
    print(f"  {'#':<4}{'Dataset':<35}{'Category':<15}{'Title'}")
    print(f"  {'─'*4}{'─'*35}{'─'*15}{'─'*40}")

    for i, name in enumerate(sorted(names), 1):
        info = TARGET_DATASETS.get(name, {})
        cat = info.get("category", "")
        title = info.get("title", name)
        marker = " *" if info else ""
        print(f"  {i:<4}{name:<35}{cat:<15}{title[:40]}{marker}")

    print(f"\n  * = in target list (auto-downloaded)")
    print()


def cmd_status(args):
    """Quick status overview."""
    if not DB_PATH.exists():
        print(f"  justice.db: NOT FOUND")
        print(f"  Run 'python justice_scraper.py download' then 'index'")
        return

    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    mtime = datetime.fromtimestamp(DB_PATH.stat().st_mtime)
    age_days = (datetime.now() - mtime).days

    conn = sqlite3.connect(str(DB_PATH))
    total = conn.execute("SELECT COUNT(*) FROM justice_data").fetchone()[0]
    n_ds = conn.execute("SELECT COUNT(DISTINCT dataset) FROM justice_data").fetchone()[0]

    cats = conn.execute(
        "SELECT category, COUNT(*), COUNT(DISTINCT dataset) FROM justice_data GROUP BY category"
    ).fetchall()
    conn.close()

    print(f"  justice.db  {db_size:.1f} MB  ({age_days}d old)")
    print(f"    observations:   {total:>10,}")
    print(f"    datasets:       {n_ds:>10}")
    for cat, count, n_ds in cats:
        label = cat or "(uncategorized)"
        print(f"      {label:<20} {count:>8,} rows  ({n_ds} datasets)")
    print()


def cmd_stats(args):
    """Show summary statistics."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))

    print(f"\n{'='*70}")
    print(f"  Justice Data — Statistics")
    print(f"{'='*70}")

    # By dataset
    print(f"\n  By Dataset:")
    for name, cat, count in conn.execute("""
        SELECT dataset, category, COUNT(*) FROM justice_data
        GROUP BY dataset ORDER BY COUNT(*) DESC
    """).fetchall():
        print(f"    {name:<35} {cat or '':<15} {count:>8,}")

    # Corruption cases (if available)
    cc_count = conn.execute("SELECT COUNT(*) FROM corruption_cases").fetchone()[0]
    if cc_count:
        print(f"\n  Corruption Cases ({cc_count:,} observations):")
        for dataset, year, value in conn.execute("""
            SELECT dataset, year, SUM(value) as total
            FROM corruption_cases WHERE year IS NOT NULL AND value IS NOT NULL
            GROUP BY dataset, year ORDER BY dataset, year
        """).fetchall():
            val_str = f"{value:,.0f}" if value else "N/A"
            print(f"    [{dataset}] {year}: {val_str}")

    # Court movements
    cm_count = conn.execute("SELECT COUNT(*) FROM court_movements").fetchone()[0]
    if cm_count:
        print(f"\n  Court Movements ({cm_count:,} observations):")
        for year, entered, finalized, pending in conn.execute("""
            SELECT year, SUM(entered), SUM(finalized), SUM(pending)
            FROM court_movements WHERE year IS NOT NULL
            GROUP BY year ORDER BY year
        """).fetchall():
            e = f"{entered:,}" if entered else "N/A"
            f_ = f"{finalized:,}" if finalized else "N/A"
            p = f"{pending:,}" if pending else "N/A"
            print(f"    {year}: Entered={e}  Finalized={f_}  Pending={p}")

    # Prison population
    pp_count = conn.execute("SELECT COUNT(*) FROM prison_population").fetchone()[0]
    if pp_count:
        print(f"\n  Prison Population ({pp_count:,} observations):")
        for year, sex, count in conn.execute("""
            SELECT year, sex, count FROM prison_population
            WHERE year IS NOT NULL ORDER BY year, sex
        """).fetchall():
            count_str = f"{count:,}" if count else "N/A"
            print(f"    {year}: {sex or 'Total'} = {count_str}")

    print(f"\n{'='*70}\n")
    conn.close()


def cmd_corruption(args):
    """Corruption case analysis."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))
    top_n = getattr(args, "top", 10)

    print(f"\n{'='*70}")
    print(f"  Corruption Cases — Analysis")
    print(f"{'='*70}")

    # Overview
    total = conn.execute("SELECT COUNT(*) FROM corruption_cases").fetchone()[0]
    if total == 0:
        print("\n  No corruption case data found.")
        print("  Run 'python justice_scraper.py download' then 'index' first.")
        conn.close()
        return

    # By dataset (corruption vs money laundering)
    print(f"\n  By Dataset:")
    for dataset, count in conn.execute("""
        SELECT dataset, COUNT(*) FROM corruption_cases GROUP BY dataset
    """).fetchall():
        print(f"    {dataset:<35} {count:>8,}")

    # Trend over years
    print(f"\n  Annual Trend:")
    for dataset in ["corrupcaopj", "branqueamentopj"]:
        rows = conn.execute("""
            SELECT year, SUM(value) as total FROM corruption_cases
            WHERE dataset = ? AND year IS NOT NULL AND value IS NOT NULL
            GROUP BY year ORDER BY year
        """, (dataset,)).fetchall()
        if rows:
            print(f"\n    {dataset}:")
            for year, total_val in rows:
                val_str = f"{total_val:,.0f}" if total_val else "N/A"
                print(f"      {year}: {val_str}")

    # By dimension (if available)
    print(f"\n  By Category/Dimension (top {top_n}):")
    for dim, count in conn.execute("""
        SELECT dimension, COUNT(*) FROM corruption_cases
        WHERE dimension != '' GROUP BY dimension ORDER BY COUNT(*) DESC LIMIT ?
    """, (top_n,)).fetchall():
        print(f"    {dim[:50]:<50} {count:>8,}")

    print(f"\n{'='*70}")
    conn.close()


def cmd_courts(args):
    """Court case flow analysis."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))

    print(f"\n{'='*70}")
    print(f"  Court Case Flow — Analysis")
    print(f"{'='*70}")

    total = conn.execute("SELECT COUNT(*) FROM court_movements").fetchone()[0]
    if total == 0:
        print("\n  No court movement data found.")
        print("  Run 'python justice_scraper.py download' then 'index' first.")
        conn.close()
        return

    # Annual summary
    print(f"\n  Annual Case Flow:")
    for year, entered, finalized, pending in conn.execute("""
        SELECT year, SUM(entered), SUM(finalized), SUM(pending)
        FROM court_movements WHERE year IS NOT NULL
        GROUP BY year ORDER BY year
    """).fetchall():
        e = f"{entered:,}" if entered else "N/A"
        f_ = f"{finalized:,}" if finalized else "N/A"
        p = f"{pending:,}" if pending else "N/A"
        resolution = ""
        if entered and finalized:
            ratio = finalized / entered * 100
            resolution = f" (resolution: {ratio:.0f}%)"
        print(f"    {year}: Entered={e}  Finalized={f_}  Pending={p}{resolution}")

    # By court type
    print(f"\n  By Court Type:")
    for court, entered, finalized in conn.execute("""
        SELECT court_type, SUM(entered), SUM(finalized)
        FROM court_movements WHERE court_type != '' AND year IS NOT NULL
        GROUP BY court_type ORDER BY SUM(entered) DESC LIMIT 10
    """).fetchall():
        e = f"{entered:,}" if entered else "N/A"
        f_ = f"{finalized:,}" if finalized else "N/A"
        print(f"    {court[:40]:<40} Entered={e}  Finalized={f_}")

    # Average duration
    dur_count = conn.execute(
        "SELECT COUNT(*) FROM justice_data WHERE dataset = 'duracaomedia'"
    ).fetchone()[0]
    if dur_count:
        print(f"\n  Average Case Duration (from duracaomedia dataset):")
        # Show raw data samples
        for row_data in conn.execute("""
            SELECT row_data FROM justice_data WHERE dataset = 'duracaomedia'
            LIMIT 5
        """).fetchall():
            data = json.loads(row_data[0])
            print(f"    {json.dumps(data, ensure_ascii=False)[:100]}")

    print(f"\n{'='*70}")
    conn.close()


def cmd_prisons(args):
    """Prison population analysis."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))

    print(f"\n{'='*70}")
    print(f"  Prison Population — Analysis")
    print(f"{'='*70}")

    total = conn.execute("SELECT COUNT(*) FROM prison_population").fetchone()[0]
    if total == 0:
        print("\n  No prison population data found.")
        print("  Run 'python justice_scraper.py download' then 'index' first.")
        conn.close()
        return

    # Annual trend
    print(f"\n  Annual Prison Population:")
    for year, sex, count in conn.execute("""
        SELECT year, sex, count FROM prison_population
        WHERE year IS NOT NULL ORDER BY year, sex
    """).fetchall():
        count_str = f"{count:,}" if count else "N/A"
        print(f"    {year}: {sex or 'Total':<10} {count_str}")

    print(f"\n{'='*70}")
    conn.close()


def cmd_query(args):
    """Run arbitrary SQL query."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    try:
        rows = conn.execute(args.sql).fetchall()
        if rows:
            headers = rows[0].keys()
            widths = [max(len(str(h)), max(len(str(r[h])) for r in rows[:50])) for h in headers]
            header_line = " | ".join(f"{h:<{w}}" for h, w in zip(headers, widths))
            print(header_line)
            print("-" * len(header_line))
            for row in rows[:100]:
                print(" | ".join(f"{str(row[h]):<{w}}" for h, w in zip(headers, widths)))
            if len(rows) > 100:
                print(f"... ({len(rows) - 100} more rows)")
        else:
            print("  (no results)")
    except Exception as e:
        print(f"  SQL Error: {e}")

    conn.close()


def cmd_export(args):
    """Export data to JSON."""
    if not DB_PATH.exists():
        print("  Database not found. Run 'download' then 'index' first.")
        return

    conn = sqlite3.connect(str(DB_PATH))

    data = {}
    rows = conn.execute("""
        SELECT dataset, category, row_data FROM justice_data
        ORDER BY dataset
    """).fetchall()

    for dataset, category, row_data in rows:
        if dataset not in data:
            data[dataset] = {"category": category, "observations": []}
        data[dataset]["observations"].append(json.loads(row_data))

    out_path = Path(args.out) if args.out else DATA_DIR / "justice_export.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"  Exported {len(rows):,} observations across {len(data)} datasets")
    print(f"  To: {out_path}")

    conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Justice Data Scraper — Portuguese court/justice statistics via CKAN API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    sub = parser.add_subparsers(dest="command")

    # List
    sub.add_parser("list", help="List all available datasets")

    # Download
    dl = sub.add_parser("download", help="Download justice datasets")
    dl.add_argument("--all", action="store_true", help="Download all 62 datasets")
    dl.add_argument("--dataset", help="Download specific dataset by name")

    # Index
    idx = sub.add_parser("index", help="Parse downloaded files into SQLite")
    idx.add_argument("--force", action="store_true", help="Re-index from scratch")

    # Status
    sub.add_parser("status", help="Quick status overview")

    # Stats
    sub.add_parser("stats", help="Summary statistics")

    # Analysis
    crp = sub.add_parser("corruption", help="Corruption case analysis")
    crp.add_argument("--top", type=int, default=10)

    crt = sub.add_parser("courts", help="Court case flow analysis")

    pri = sub.add_parser("prisons", help="Prison population analysis")

    # Query
    qr = sub.add_parser("query", help="Run arbitrary SQL")
    qr.add_argument("sql", help="SQL query to execute")

    # Export
    exp = sub.add_parser("export", help="Export to JSON")
    exp.add_argument("--out", help="Output file path")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "list": cmd_list,
        "download": cmd_download,
        "index": cmd_index,
        "status": cmd_status,
        "stats": cmd_stats,
        "corruption": cmd_corruption,
        "courts": cmd_courts,
        "prisons": cmd_prisons,
        "query": cmd_query,
        "export": cmd_export,
    }

    commands[args.command](args)


if __name__ == "__main__":
    main()
