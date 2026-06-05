#!/usr/bin/env python3
"""Merge NIFs from dados.gov.pt contratos datasets into bep_entities index.

Downloads the IMPIC public procurement dataset (contratos2025.xlsx) from
dados.gov.pt, extracts NIF + entity name pairs from the 'adjudicante' column,
and matches them against our bep_entities by fuzzy name matching.

Usage:
    python merge_nifs.py                  # Run merge with defaults
    python merge_nifs.py --dry-run        # Preview matches without writing
    python merge_nifs.py --threshold 80   # Lower fuzzy match threshold (default 75)
"""

import sqlite3
import re
import sys
import os
import subprocess
import unicodedata
from pathlib import Path
from difflib import SequenceMatcher

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DB_PATH = SCRIPT_DIR / "bep_index.db"
DATA_DIR = SCRIPT_DIR / "data"
XLSX_PATH = DATA_DIR / "contratos2025.xlsx"

# dados.gov.pt dataset URL for contratos 2025
CONTRATOS_URL = "https://dados.gov.pt/s/resources/contratos-publicos-portal-base-impic-contratos-de-2012-a-2026/20260601-125604/contratos2025.xlsx"


def normalize_name(name: str) -> str:
    """Normalize entity name for fuzzy matching."""
    name = name.lower().strip()
    # Remove accents
    name = ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    # Remove common suffixes
    for suffix in [', i.p.', ', ip', ', i. p.', ', p.u.', ', p. u.', ', s.a.', ' - sede', ' (sede)']:
        name = name.replace(suffix, '')
    # Normalize whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def extract_nif_from_adjudicante(text: str) -> tuple[str, str]:
    """Extract NIF and entity name from adjudicante field.
    
    Format: 'NIF - Entity Name' or '- - Entity Name' (no NIF)
    Returns: (nif, entity_name)
    """
    if not text:
        return ("", "")
    
    # Pattern: "123456789 - Entity Name"  
    m = re.match(r'^(\d{9})\s*-\s*(.+)$', text.strip())
    if m:
        return (m.group(1), m.group(2).strip())
    
    # Pattern: "- - Entity Name" (no NIF)
    m = re.match(r'^-\s*-\s*(.+)$', text.strip())
    if m:
        return ("", m.group(1).strip())
    
    return ("", text.strip())


def download_xlsx(url: str, dest: Path) -> bool:
    """Download the XLSX file if not already present."""
    if dest.exists() and dest.stat().st_size > 1000:
        print(f"  Using cached: {dest.name} ({dest.stat().st_size / 1024 / 1024:.1f} MB)")
        return True
    
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Downloading from dados.gov.pt...")
    try:
        result = subprocess.run(
            ["curl", "-s", "-L", "-o", str(dest), url],
            capture_output=True, timeout=120,
        )
        if dest.exists() and dest.stat().st_size > 1000:
            print(f"  Downloaded: {dest.stat().st_size / 1024 / 1024:.1f} MB")
            return True
        print(f"  ERROR: Download failed or file too small")
        return False
    except Exception as e:
        print(f"  ERROR: {e}")
        return False


def build_nif_lookup_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    """Parse the XLSX and build a lookup: normalized_entity_name -> nif.
    
    Only includes entities with valid NIFs (9-digit numbers).
    """
    import openpyxl
    
    print(f"  Parsing {xlsx_path.name}...")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Find the adjudicante column
    headers = next(ws.iter_rows(max_row=1, values_only=True))
    adj_idx = None
    for i, h in enumerate(headers):
        if h and 'adjudicante' in str(h).lower():
            adj_idx = i
            break
    
    if adj_idx is None:
        print("  ERROR: 'adjudicante' column not found!")
        wb.close()
        return {}    # Extract NIF + name pairs
    name_to_nif: dict[str, str] = {}
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        text = str(row[adj_idx]) if row[adj_idx] else ""
        nif, name = extract_nif_from_adjudicante(text)

        if nif and name:
            norm = normalize_name(name)
            if norm not in name_to_nif:
                name_to_nif[norm] = nif

    wb.close()
    print(f"  Parsed {total} contracts → {len(name_to_nif)} unique entities with NIFs")
    return name_to_nif


def fuzzy_match(query: str, candidates: dict[str, str], threshold: float = 0.75) -> tuple[str, float]:
    """Find the best fuzzy match for query in candidates.
    
    Returns: (matched_key, score) or ("", 0.0)
    """
    q = normalize_name(query)
    best_key = ""
    best_score = 0.0
    
    for key in candidates:
        # Exact substring match is best
        if q in key or key in q:
            score = 0.95
        else:
            score = SequenceMatcher(None, q, key).ratio()
        
        if score > best_score:
            best_score = score
            best_key = key
    
    return (best_key, best_score)


def merge_nifs(dry_run: bool = False, threshold: float = 0.75):
    """Main merge logic."""
    # Init DB
    sys.path.insert(0, str(SCRIPT_DIR))
    from bep_db import init_db, set_nif, get_entities_without_nif, search_entities
    
    conn = init_db(str(DB_PATH))
    
    # Get entities without NIF
    entities = get_entities_without_nif(conn)
    if not entities:
        print("All entities already have NIFs!")
        conn.close()
        return
    
    print(f"\nEntities without NIF: {len(entities)}")
    
    # Download and parse XLSX
    if not download_xlsx(CONTRATOS_URL, XLSX_PATH):
        print("ERROR: Could not download contratos dataset")
        conn.close()
        return
    
    name_to_nif = build_nif_lookup_from_xlsx(XLSX_PATH)
    if not name_to_nif:
        print("ERROR: No NIFs extracted from dataset")
        conn.close()
        return
    
    # Match entities against NIF lookup
    matched = 0
    unmatched = []
    
    for e in entities:
        display = e['organismo'] or e['entidade']
        match_key, score = fuzzy_match(display, name_to_nif, threshold)
        
        if score >= threshold and match_key:
            nif = name_to_nif[match_key]
            if dry_run:
                print(f"  [DRY-RUN] {e['id']}  NIF={nif}  {display[:50]:50s}  ← {match_key[:50]} ({score:.2f})")
            else:
                set_nif(conn, e['id'], nif)
                print(f"  ✓ {e['id']}  NIF={nif}  {display[:50]:50s}  ← {match_key[:50]} ({score:.2f})")
            matched += 1
        else:
            unmatched.append((e, score))
    
    if not dry_run:
        conn.commit()
    
    conn.close()
    
    # Summary
    print(f"\n=== Merge complete ===")
    print(f"Matched: {matched}/{len(entities)} entities got NIFs")
    if unmatched:
        print(f"Unmatched ({len(unmatched)}):")
        for e, score in unmatched[:10]:
            display = e['organismo'] or e['entidade']
            print(f"  {e['id']}  (best score: {score:.2f})  {display[:70]}")
        if len(unmatched) > 10:
            print(f"  ... and {len(unmatched) - 10} more")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Merge NIFs from dados.gov.pt into BEP entity index")
    parser.add_argument("--dry-run", action="store_true", help="Preview matches without writing")
    parser.add_argument("--threshold", type=float, default=0.75, help="Fuzzy match threshold (0-1)")
    args = parser.parse_args()
    
    merge_nifs(dry_run=args.dry_run, threshold=args.threshold)
