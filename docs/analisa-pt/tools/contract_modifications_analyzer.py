#!/usr/bin/env python3
"""Contract Modifications Analyzer — Detect suspicious contract amendments.

Downloads and indexes 15 years (2012–2026) of modificações contratuais from
dados.gov.pt, cross-references with contract_index.json, and detects corruption
signals: excessive price inflation, frequent amendments, unilateral changes,
and 'acordo entre as partes' abuse.

Usage:
    python contract_modifications_analyzer.py download    # Download all XLSX
    python contract_modifications_analyzer.py index       # Parse → SQLite
    python contract_modifications_analyzer.py stats       # Summary stats
    python contract_modifications_analyzer.py suspicious  # Flag suspicious patterns
    python contract_modifications_analyzer.py inflation   # Price inflation analysis
    python contract_modifications_analyzer.py frequent    # Frequently amended contracts
    python contract_modifications_analyzer.py justification # Analyze justification types
    python contract_modifications_analyzer.py entity --nif X  # Entity-specific
    python contract_modifications_analyzer.py crossref    # Cross-ref with contracts
    python contract_modifications_analyzer.py export      # Export to JSON
"""

import sys
import json
import re
import sqlite3
import argparse
import time
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone

try:
    import urllib.request
    import ssl
except ImportError:
    print("ERROR: urllib required (built-in)")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"
DB_PATH = DATA_DIR / "modificacoes_index.db"
CONTRACT_INDEX = DATA_DIR / "contract_index.json"
XLSX_DIR = DATA_DIR

# SSL context
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/octet-stream,*/*",
}

# dados.gov.pt dataset ID for modificações contratuais
DATASET_ID = "668d65dbcb1b953e80198435"
API_BASE = "https://dados.gov.pt/api/1/datasets/"


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def init_db() -> sqlite3.Connection:
    """Initialize the modifications database."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS modificacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            idcontrato INTEGER,
            fundamento TEXT,
            tipo_acto TEXT,
            data_modificacao TEXT,
            preco_alterado REAL,
            prazo_execucao INTEGER,
            data_publicacao TEXT,
            ano INTEGER,
            UNIQUE(idcontrato, fundamento, data_modificacao)
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS download_log (
            year INTEGER PRIMARY KEY,
            filename TEXT,
            size_bytes INTEGER,
            rows INTEGER,
            downloaded_at TEXT,
            indexed_at TEXT
        )
    """)

    conn.execute("CREATE INDEX IF NOT EXISTS idx_mod_idcontrato ON modificacoes(idcontrato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mod_fundamento ON modificacoes(fundamento)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mod_tipo ON modificacoes(tipo_acto)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mod_ano ON modificacoes(ano)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mod_preco ON modificacoes(preco_alterado)")

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def get_download_urls() -> dict[int, str]:
    """Fetch per-year download URLs from the dados.gov.pt API."""
    url = f"{API_BASE}{DATASET_ID}"
    req = urllib.request.Request(url, headers=HEADERS)
    resp = urllib.request.urlopen(req, timeout=15, context=SSL_CTX)
    data = json.loads(resp.read())

    ds = data.get("data", data)
    resources = ds.get("resources", [])

    urls = {}
    for r in resources:
        name = r.get("title", r.get("name", ""))
        fmt = str(r.get("format", "")).lower()
        dl_url = r.get("url", "")
        if fmt == "xlsx" and name.startswith("modcontra"):
            m = re.search(r"(\d{4})", name)
            if m:
                urls[int(m.group(1))] = dl_url

    return urls


def cmd_download(args):
    """Download all modificações contratuais XLSX files."""
    print("Fetching download URLs from dados.gov.pt API...")
    urls = get_download_urls()
    print(f"  Found {len(urls)} XLSX files (years {min(urls)}-{max(urls)})\n")

    conn = init_db()
    total_size = 0

    for year in sorted(urls.keys()):
        xlsx_path = XLSX_DIR / f"modcontra{year}.xlsx"

        if xlsx_path.exists() and xlsx_path.stat().st_size > 1000:
            size = xlsx_path.stat().st_size
            print(f"  {year}: already exists ({size:,} bytes)")
            total_size += size
            continue

        print(f"  {year}: downloading...", end=" ", flush=True)
        try:
            req = urllib.request.Request(urls[year], headers=HEADERS)
            resp = urllib.request.urlopen(req, timeout=60, context=SSL_CTX)
            data = resp.read()
            xlsx_path.write_bytes(data)
            size = len(data)
            total_size += size
            print(f"done ({size:,} bytes)")
        except Exception as e:
            print(f"FAILED: {e}")
            continue

        time.sleep(0.5)

    print(f"\n  Total: {total_size:,} bytes across {len(urls)} files")

    for year, url in urls.items():
        xlsx_path = XLSX_DIR / f"modcontra{year}.xlsx"
        if xlsx_path.exists():
            conn.execute(
                "INSERT OR REPLACE INTO download_log (year, filename, size_bytes, downloaded_at) "
                "VALUES (?, ?, ?, ?)",
                (year, xlsx_path.name, xlsx_path.stat().st_size,
                 datetime.now(timezone.utc).isoformat()),
            )
    conn.commit()
    conn.close()


# ---------------------------------------------------------------------------
# Index (XLSX → SQLite)
# ---------------------------------------------------------------------------

def parse_xlsx(year: int) -> list[tuple]:
    """Parse a single year's XLSX file into rows."""
    xlsx_path = XLSX_DIR / f"modcontra{year}.xlsx"
    if not xlsx_path.exists():
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        def get(field, default=""):
            idx = header_map.get(field)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return val if val is not None else default

        def get_num(field):
            val = get(field, None)
            if val is None:
                return None
            try:
                v = float(val)
                # Cap at €10 billion — values above are data entry errors
                if v > 1e10:
                    return None
                return v if v > 0 else None
            except (ValueError, TypeError):
                return None

        def fmt_date(val):
            if val is None:
                return ""
            if hasattr(val, 'strftime'):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()[:10]

        def get_int(field):
            val = get(field, None)
            if val is None:
                return None
            try:
                return int(val)
            except (ValueError, TypeError):
                return None

        idcontrato = get_int("idcontrato")
        if not idcontrato:
            continue

        rows.append((
            idcontrato,
            str(get("modifContratoFundamento", "")).strip()[:500],
            str(get("modifContratoTipoAto", "")).strip(),
            fmt_date(get("modifContratoData")),
            get_num("modifContratoPrecoAlterado"),
            get_int("modifPrazoExecucao"),
            fmt_date(get("modifDataPublicacao")),
            year,
        ))

    wb.close()
    return rows


def cmd_index(args):
    """Parse all downloaded XLSX files into SQLite."""
    conn = init_db()

    if args.force:
        conn.execute("DELETE FROM modificacoes")
        conn.commit()
        print("Cleared existing index.\n")

    years = range(2012, 2027)
    total_rows = 0
    total_new = 0

    for year in years:
        xlsx_path = XLSX_DIR / f"modcontra{year}.xlsx"
        if not xlsx_path.exists():
            print(f"  {year}: file not found, skipping")
            continue

        print(f"  {year}: parsing...", end=" ", flush=True)
        rows = parse_xlsx(year)

        changes_before = conn.total_changes
        conn.executemany(
            "INSERT OR IGNORE INTO modificacoes "
            "(idcontrato, fundamento, tipo_acto, data_modificacao, preco_alterado, "
            "prazo_execucao, data_publicacao, ano) "
            "VALUES (?,?,?,?,?,?,?,?)",
            rows,
        )
        new_count = conn.total_changes - changes_before
        total_new += new_count
        total_rows += len(rows)

        conn.execute(
            "UPDATE download_log SET rows=?, indexed_at=? WHERE year=?",
            (len(rows), datetime.now(timezone.utc).isoformat(), year),
        )

        print(f"{len(rows):,} rows ({new_count:,} new)")
        conn.commit()

    print(f"\n  Total: {total_rows:,} rows parsed, {total_new:,} inserted")
    count = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    print(f"  Index total: {count:,} modifications")
    conn.close()


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def cmd_stats(args):
    """Show summary statistics."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Modificações Contratuais — Statistics")
    print(f"{'='*70}")
    print(f"\n  Total modifications: {total:,}\n")

    # By year
    print(f"  By Year:")
    year_data = conn.execute(
        "SELECT ano, COUNT(*), COUNT(DISTINCT idcontrato), "
        "SUM(COALESCE(preco_alterado, 0)) "
        "FROM modificacoes GROUP BY ano ORDER BY ano"
    ).fetchall()
    for year, count, contracts, total_val in year_data:
        avg = (total_val / count) if count and total_val else 0
        print(f"    {year}: {count:>6,} modifications  "
              f"({contracts:>5,} contracts)  "
              f"€{total_val:>14,.0f}  (avg €{avg:>10,.0f})")

    # Justification types
    print(f"\n  Justification Types (tipo_acto):")
    tipo_data = conn.execute(
        "SELECT tipo_acto, COUNT(*), SUM(COALESCE(preco_alterado, 0)) "
        "FROM modificacoes WHERE tipo_acto != '' "
        "GROUP BY tipo_acto ORDER BY COUNT(*) DESC"
    ).fetchall()
    for tipo, count, total_val in tipo_data:
        pct = count * 100 / total
        print(f"    {tipo[:55]:55s} {count:>6,} ({pct:>5.1f}%)  €{total_val:>14,.0f}")

    # Top justification grounds
    print(f"\n  Top Justification Grounds (fundamento):")
    fund_data = conn.execute(
        "SELECT fundamento, COUNT(*), AVG(preco_alterado) "
        "FROM modificacoes WHERE fundamento != '' "
        "GROUP BY fundamento ORDER BY COUNT(*) DESC LIMIT 15"
    ).fetchall()
    for fund, count, avg_val in fund_data:
        avg_str = f"€{avg_val:,.0f}" if avg_val else "N/A"
        print(f"    {fund[:55]:55s} {count:>6,}  avg {avg_str}")

    print(f"{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Entity name resolution
# ---------------------------------------------------------------------------

def _load_entity_lookup() -> dict[int, str]:
    """Load contract_id → entity_name from contract_index.json."""
    if not CONTRACT_INDEX.exists():
        return {}
    with open(CONTRACT_INDEX, "r", encoding="utf-8") as f:
        idx = json.load(f)
    lookup = {}
    for nif, contracts in idx.items():
        for c in contracts:
            cid = c.get("contract_id", 0)
            if cid:
                lookup[cid] = c.get("entity_name", nif)
    return lookup


def _entity_name(lookup: dict[int, str], cid: int) -> str:
    """Resolve contract ID to entity name."""
    return lookup.get(cid, f"Contract #{cid}")


# ---------------------------------------------------------------------------
# Suspicious patterns
# ---------------------------------------------------------------------------

def cmd_suspicious(args):
    """Flag suspicious modification patterns."""
    conn = init_db()
    lookup = _load_entity_lookup()
    ename = lambda cid: _entity_name(lookup, cid)

    total = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Modificações Contratuais — Suspicious Patterns")
    print(f"{'='*70}\n")

    # 1. High-value modifications (single amendment > €1M)
    print(f"  🔴 HIGH-VALUE MODIFICATIONS (>€1M single amendment):")
    high_val = conn.execute(
        "SELECT idcontrato, fundamento, tipo_acto, preco_alterado, data_modificacao "
        "FROM modificacoes WHERE preco_alterado > 1000000 "
        "ORDER BY preco_alterado DESC LIMIT 20"
    ).fetchall()
    for cid, fund, tipo, preco, data in high_val:
        print(f"    {ename(cid)[:50]}")
        print(f"      €{preco:>14,.0f}  [{tipo[:30]}]  {data}")
        print(f"      {fund[:70]}")
        print()

    # 2. 'Acordo entre as partes' (bilateral agreement) — often circumvents competition
    ad_count = conn.execute(
        "SELECT COUNT(*) FROM modificacoes WHERE tipo_acto LIKE '%Acordo entre as partes%'"
    ).fetchone()[0]
    total_with_tipo = conn.execute(
        "SELECT COUNT(*) FROM modificacoes WHERE tipo_acto != ''"
    ).fetchone()[0]
    print(f"  🟡 'ACORDO ENTRE AS PARTES' (bilateral agreement):")
    print(f"    Count: {ad_count:,} / {total_with_tipo:,} ({ad_count*100/total_with_tipo:.1f}%)")
    print(f"    These bypass competitive procedures under CCP art. 370/454\n")

    # 3. Frequent amendments (same contract modified 3+ times)
    print(f"  🟠 FREQUENT AMENDMENTS (3+ modifications per contract):")
    frequent = conn.execute(
        "SELECT idcontrato, COUNT(*) as cnt, SUM(COALESCE(preco_alterado, 0)) as total_val "
        "FROM modificacoes GROUP BY idcontrato HAVING cnt >= 3 "
        "ORDER BY cnt DESC LIMIT 20"
    ).fetchall()
    for cid, cnt, total_val in frequent:
        print(f"    {ename(cid)[:50]}: {cnt} mods, €{total_val:,.0f}")

    # 4. Scope creep (justification mentions multiple CCP articles)
    print(f"\n  🟡 SCOPE CREEP (multiple CCP articles in justification):")
    scope_creo = conn.execute(
        "SELECT idcontrato, fundamento, COUNT(*) "
        "FROM modificacoes WHERE fundamento LIKE '%art.%' OR fundamento LIKE '%artigo%' "
        "GROUP BY idcontrato HAVING COUNT(*) >= 2 "
        "ORDER BY COUNT(*) DESC LIMIT 10"
    ).fetchall()
    for cid, fund, cnt in scope_creo:
        print(f"    {ename(cid)[:50]}: {cnt} amendments")
        print(f"      {fund[:70]}")

    # 5. Long-deadline modifications (>1000 days execution period)
    print(f"\n  🔴 LONG DEADLINE MODIFICATIONS (>1000 days execution period):")
    long_deadline = conn.execute(
        "SELECT idcontrato, prazo_execucao, data_modificacao, preco_alterado, fundamento "
        "FROM modificacoes WHERE prazo_execucao > 1000 "
        "ORDER BY prazo_execucao DESC LIMIT 10"
    ).fetchall()
    for cid, prazo, data_mod, preco, fund in long_deadline:
        preco_str = f"€{preco:,.0f}" if preco else "N/A"
        print(f"    {ename(cid)[:50]}")
        print(f"      {prazo} days, {data_mod}, {preco_str}")
        print(f"      {fund[:60]}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Price inflation
# ---------------------------------------------------------------------------

def cmd_inflation(args):
    """Analyze price inflation through modifications."""
    conn = init_db()
    lookup = _load_entity_lookup()
    ename = lambda cid: _entity_name(lookup, cid)

    total = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Modificações — Price Inflation Analysis")
    print(f"{'='*70}\n")

    # Contracts with multiple price increases
    print(f"  🔴 CUMULATIVE PRICE INFLATION (contracts with multiple price increases):")
    inflation = conn.execute(
        "SELECT idcontrato, COUNT(*) as cnt, "
        "SUM(COALESCE(preco_alterado, 0)) as total_increase, "
        "AVG(preco_alterado) as avg_increase "
        "FROM modificacoes WHERE preco_alterado > 0 "
        "GROUP BY idcontrato HAVING cnt >= 2 AND total_increase > 100000 "
        "ORDER BY total_increase DESC LIMIT 20"
    ).fetchall()
    for cid, cnt, total_inc, avg_inc in inflation:
        print(f"    {ename(cid)[:50]}")
        print(f"      {cnt} increases, cumulative +€{total_inc:,.0f} (avg +€{avg_inc:,.0f})")

    # Year-over-year modification value trends
    print(f"\n  📈 YEAR-OVER-YEAR MODIFICATION VALUES:")
    print(f"    {'Year':>6}  {'Mods':>7}  {'Contracts':>10}  {'Total Value':>15}  {'Avg Value':>12}")
    print(f"    {'─'*6}  {'─'*7}  {'─'*10}  {'─'*15}  {'─'*12}")
    year_data = conn.execute(
        "SELECT ano, COUNT(*), COUNT(DISTINCT idcontrato), "
        "SUM(COALESCE(preco_alterado, 0)), "
        "AVG(CASE WHEN preco_alterado > 0 THEN preco_alterado END) "
        "FROM modificacoes GROUP BY ano ORDER BY ano"
    ).fetchall()
    for year, count, contracts, total_val, avg_val in year_data:
        print(f"    {year:>6}  {count:>7,}  {contracts:>10,}  €{total_val:>14,.0f}  €{avg_val:>11,.0f}")

    # Modification justification breakdown by year
    print(f"\n  JUSTIFICATION TYPE EVOLUTION:")
    tipo_year = conn.execute(
        "SELECT ano, tipo_acto, COUNT(*) FROM modificacoes "
        "WHERE tipo_acto != '' GROUP BY ano, tipo_acto ORDER BY ano, COUNT(*) DESC"
    ).fetchall()
    tipo_by_year = defaultdict(dict)
    for year, tipo, count in tipo_year:
        tipo_by_year[year][tipo] = count
    for year in sorted(tipo_by_year.keys()):
        top = sorted(tipo_by_year[year].items(), key=lambda x: -x[1])[:3]
        summary = ", ".join(f"{t}: {c}" for t, c in top)
        print(f"    {year}: {summary}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Frequent amendments
# ---------------------------------------------------------------------------

def cmd_frequent(args):
    """Find frequently amended contracts."""
    conn = init_db()
    lookup = _load_entity_lookup()
    ename = lambda cid: _entity_name(lookup, cid)

    total = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Modificações — Frequently Amended Contracts")
    print(f"{'='*70}\n")

    # Top 30 most amended contracts
    print(f"  Top 30 Most Amended Contracts:")
    frequent = conn.execute(
        "SELECT idcontrato, COUNT(*) as cnt, "
        "MIN(data_modificacao), MAX(data_modificacao), "
        "SUM(COALESCE(preco_alterado, 0)), "
        "GROUP_CONCAT(DISTINCT tipo_acto) "
        "FROM modificacoes GROUP BY idcontrato "
        "ORDER BY cnt DESC LIMIT 30"
    ).fetchall()
    for cid, cnt, first, last, val, tipos in frequent:
        tipos_short = (tipos or "")[:45]
        val_str = f"€{val:,.0f}" if val else "N/A"
        print(f"  {ename(cid)[:45]:45s}  {cnt:>3} mods  {val_str:>14}  {first or 'N/A'}→{last or 'N/A'}")

    # Contracts modified in multiple years (long-term scope creep)
    print(f"\n  🔴 LONG-TERM SCOPE CREEP (modified across 3+ different years):")
    multi_year = conn.execute(
        "SELECT idcontrato, COUNT(DISTINCT ano) as years, COUNT(*) as mods, "
        "SUM(COALESCE(preco_alterado, 0)) "
        "FROM modificacoes GROUP BY idcontrato HAVING years >= 3 "
        "ORDER BY years DESC, mods DESC LIMIT 15"
    ).fetchall()
    for cid, years, mods, val in multi_year:
        val_str = f"€{val:,.0f}" if val else "N/A"
        print(f"    {ename(cid)[:50]}: {years} years, {mods} mods, {val_str}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Justification analysis
# ---------------------------------------------------------------------------

def cmd_justification(args):
    """Analyze modification justifications."""
    conn = init_db()

    total = conn.execute("SELECT COUNT(*) FROM modificacoes").fetchone()[0]
    if total == 0:
        print("  Index is empty. Run 'index' first.")
        conn.close()
        return

    print(f"\n{'='*70}")
    print(f"  Modificações — Justification Analysis")
    print(f"{'='*70}\n")

    # CCP article frequency
    print(f"  CCP Article References in Justifications:")
    ccp_articles = conn.execute(
        "SELECT fundamento, COUNT(*) as cnt, AVG(preco_alterado) as avg_val "
        "FROM modificacoes WHERE fundamento LIKE '%art.%' "
        "GROUP BY fundamento ORDER BY cnt DESC LIMIT 20"
    ).fetchall()
    for fund, cnt, avg_val in ccp_articles:
        avg_str = f"€{avg_val:,.0f}" if avg_val else "N/A"
        print(f"    {fund[:60]:60s} {cnt:>5}  avg {avg_str}")

    # 'Acordo entre as partes' vs other types
    print(f"\n  Modification Type Comparison:")
    tipo_stats = conn.execute(
        "SELECT tipo_acto, COUNT(*), "
        "AVG(CASE WHEN preco_alterado > 0 THEN preco_alterado END), "
        "SUM(COALESCE(preco_alterado, 0)) "
        "FROM modificacoes WHERE tipo_acto != '' "
        "GROUP BY tipo_acto ORDER BY COUNT(*) DESC"
    ).fetchall()
    for tipo, cnt, avg_val, total_val in tipo_stats:
        avg_str = f"€{avg_val:,.0f}" if avg_val else "N/A"
        total_str = f"€{total_val:,.0f}" if total_val else "N/A"
        print(f"    {tipo[:50]:50s} {cnt:>6,}  avg {avg_str:>12}  total {total_str:>14}")

    print(f"\n{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Entity search
# ---------------------------------------------------------------------------

def cmd_entity(args):
    """Show modifications for a specific entity."""
    conn = init_db()

    if not CONTRACT_INDEX.exists():
        print(f"  Contract index not found at {CONTRACT_INDEX}")
        conn.close()
        return

    with open(CONTRACT_INDEX, "r", encoding="utf-8") as f:
        contract_index = json.load(f)

    # Find contracts for this entity
    nif = args.nif
    entity_contracts = set()
    entity_name = nif
    for n, contracts in contract_index.items():
        if n == nif:
            entity_name = contracts[0].get("entity_name", nif) if contracts else nif
            for c in contracts:
                entity_contracts.add(c.get("contract_id", 0))
            break

    if not entity_contracts:
        print(f"  No contracts found for NIF {nif}")
        conn.close()
        return

    mods = conn.execute(
        "SELECT idcontrato, fundamento, tipo_acto, data_modificacao, preco_alterado, prazo_execucao "
        "FROM modificacoes WHERE idcontrato IN ({}) "
        "ORDER BY data_modificacao DESC LIMIT ?".format(",".join("?" * len(entity_contracts))),
        list(entity_contracts) + [args.limit],
    ).fetchall()

    print(f"\n{'='*70}")
    print(f"  Entity: {entity_name} (NIF: {nif})")
    print(f"  Contracts in index: {len(entity_contracts):,}")
    print(f"  Modifications found: {len(mods):,}")
    print(f"{'='*70}\n")

    for cid, fund, tipo, data, preco, prazo in mods:
        preco_str = f"€{preco:,.0f}" if preco else "N/A"
        print(f"  Contract #{cid} | {data} | {tipo}")
        print(f"    Justification: {fund[:70]}")
        print(f"    Price: {preco_str} | Deadline: {prazo} days")
        print()

    conn.close()


# ---------------------------------------------------------------------------
# Cross-reference with contracts
# ---------------------------------------------------------------------------

def cmd_crossref(args):
    """Cross-reference modifications with signed contracts."""
    conn = init_db()

    if not CONTRACT_INDEX.exists():
        print(f"  Contract index not found at {CONTRACT_INDEX}")
        conn.close()
        return

    with open(CONTRACT_INDEX, "r", encoding="utf-8") as f:
        contract_index = json.load(f)

    contract_lookup = {}
    for nif, contracts in contract_index.items():
        for c in contracts:
            cid = c.get("contract_id", 0)
            if cid:
                contract_lookup[cid] = c

    # Batch query: get modification counts and sums for all contracts at once
    mod_summary = conn.execute(
        "SELECT idcontrato, COUNT(*) as cnt, SUM(COALESCE(preco_alterado, 0)) as total_val "
        "FROM modificacoes WHERE idcontrato > 0 GROUP BY idcontrato"
    ).fetchall()

    matched_ids = set(r[0] for r in mod_summary) & set(contract_lookup.keys())
    mod_map = {r[0]: (r[1], r[2]) for r in mod_summary}

    print(f"\n{'='*70}")
    print(f"  Modificações × Contracts Cross-Reference")
    print(f"{'='*70}")
    print(f"  Modifications index: {conn.execute('SELECT COUNT(*) FROM modificacoes').fetchone()[0]:,}")
    print(f"  Contracts index: {sum(len(v) for v in contract_index.values()):,}")
    print(f"  Contracts with modifications: {len(matched_ids):,}")

    if matched_ids:
        print(f"\n  Top 20 matched contracts by modification value:")
        ranked = []
        for cid in matched_ids:
            c = contract_lookup[cid]
            cnt, val = mod_map[cid]
            ranked.append((cid, c.get("entity_name", ""), c.get("valor", 0), cnt, val))

        ranked.sort(key=lambda x: -x[4])
        for cid, name, contract_val, mod_count, mod_total in ranked[:20]:
            pct = (mod_total / contract_val * 100) if contract_val > 0 else 0
            print(f"    #{cid:>10}  {name[:35]:35s}  "
                  f"contract €{contract_val:>12,.0f}  "
                  f"{mod_count} mods  +€{mod_total:>12,.0f} ({pct:.0f}%)")

    print(f"{'='*70}\n")
    conn.close()


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def cmd_export(args):
    """Export index to JSON."""
    conn = init_db()

    rows = conn.execute(
        "SELECT idcontrato, fundamento, tipo_acto, data_modificacao, "
        "preco_alterado, prazo_execucao, data_publicacao, ano "
        "FROM modificacoes ORDER BY ano, data_modificacao"
    ).fetchall()

    data = []
    for row in rows:
        data.append({
            "idcontrato": row[0],
            "fundamento": row[1],
            "tipo_acto": row[2],
            "data_modificacao": row[3],
            "preco_alterado": row[4],
            "prazo_execucao": row[5],
            "data_publicacao": row[6],
            "ano": row[7],
        })

    out_path = Path(args.out) if args.out else DATA_DIR / "modificacoes_index.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"  Exported {len(data):,} modifications to {out_path}")
    conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Contract Modifications Analyzer — Portuguese contract amendments (2012–2026)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    sub = parser.add_subparsers(dest="command")

    sub.add_parser("download", help="Download all XLSX files from dados.gov.pt")

    idx = sub.add_parser("index", help="Parse XLSX files into SQLite index")
    idx.add_argument("--force", action="store_true", help="Re-index from scratch")

    sub.add_parser("stats", help="Summary statistics")
    sub.add_parser("suspicious", help="Flag suspicious modification patterns")
    sub.add_parser("inflation", help="Price inflation analysis")
    sub.add_parser("frequent", help="Frequently amended contracts")
    sub.add_parser("justification", help="Justification type analysis")

    ent = sub.add_parser("entity", help="Entity-specific modifications")
    ent.add_argument("--nif", required=True, help="Filter by entity NIF")
    ent.add_argument("--limit", type=int, default=20)

    sub.add_parser("crossref", help="Cross-ref with contract_index.json")

    exp = sub.add_parser("export", help="Export to JSON")
    exp.add_argument("--out", help="Output file path")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "download": cmd_download,
        "index": cmd_index,
        "stats": cmd_stats,
        "suspicious": cmd_suspicious,
        "inflation": cmd_inflation,
        "frequent": cmd_frequent,
        "justification": cmd_justification,
        "entity": cmd_entity,
        "crossref": cmd_crossref,
        "export": cmd_export,
    }

    commands[args.command](args)


if __name__ == "__main__":
    main()
