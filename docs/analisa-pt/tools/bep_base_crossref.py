#!/usr/bin/env python3
"""BEP × BASE.gov.pt Cross-Reference Report

Links BEP public sector job entities (via NIF) to their public procurement
contracts from BASE.gov.pt. Identifies entities that are both hiring staff
AND awarding public contracts — a key transparency signal.

Usage:
    python bep_base_crossref.py                      # Full report
    python bep_base_crossref.py --entity "Gaia"      # Filter by entity name
    python bep_base_crossref.py --nif 500014872      # Filter by NIF
    python bep_base_crossref.py --export report.json  # Export to JSON
    python bep_base_crossref.py --stats               # Summary statistics
"""

import sys
import json
import re
import argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
DB_PATH = SCRIPT_DIR / "bep_index.db"
XLSX_PATH = SCRIPT_DIR / "data" / "contratos2025.xlsx"
BASE_DETAIL_URL = "https://www.base.gov.pt/Base4/pt/detalhe/?type=contratos&id="


def extract_nif_from_adjudicante(text: str) -> tuple[str, str]:
    """Extract NIF and entity name from adjudicante field."""
    if not text:
        return ("", "")
    m = re.match(r'^(\d{9})\s*-\s*(.+)$', text.strip())
    if m:
        return (m.group(1), m.group(2).strip())
    m = re.match(r'^-\s*-\s*(.+)$', text.strip())
    if m:
        return ("", m.group(1).strip())
    return ("", text.strip())


def load_bep_entities_with_nif(db_path: Path) -> list[dict]:
    """Load BEP entities that have NIFs."""
    import sqlite3
    conn = sqlite3.connect(str(db_path))
    rows = conn.execute(
        "SELECT id, entidade, organismo, display_name, nif, listing_count "
        "FROM bep_entities WHERE nif IS NOT NULL AND nif != '' "
        "ORDER BY listing_count DESC"
    ).fetchall()
    conn.close()
    return [
        {"id": r[0], "entidade": r[1], "organismo": r[2],
         "display_name": r[3], "nif": r[4], "listing_count": r[5]}
        for r in rows
    ]


CACHE_PATH = SCRIPT_DIR / "data" / "contract_index.json"


def build_contract_index(xlsx_path: Path) -> dict[str, list[dict]]:
    """Parse contratos XLSX and build NIF → contracts index. Caches as JSON."""
    if CACHE_PATH.exists():
        print(f"  Loading cached index from {CACHE_PATH.name}...")
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)

    print(f"  Parsing {xlsx_path.name} (this takes a while for 67MB)...\n")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = next(ws.iter_rows(max_row=1, values_only=True))
    header_map = {str(h).lower().strip(): i for i, h in enumerate(headers) if h}

    adj_idx = header_map.get("adjudicante")
    id_idx = header_map.get("idcontrato")
    link_idx = header_map.get("linkpecasproc")
    valor_idx = header_map.get("precocontratual") or header_map.get("precobaseprocedimento") or header_map.get("valor") or header_map.get("valorcontrato")
    data_idx = header_map.get("datacelebracaocontrato") or header_map.get("datapublicacao") or header_map.get("data") or header_map.get("datacontrato")
    tipo_idx = header_map.get("tipocontrato") or header_map.get("tipoprocedimento") or header_map.get("tipo")
    desc_idx = header_map.get("objectocontrato") or header_map.get("desccontrato") or header_map.get("objeto")

    if adj_idx is None:
        print("  ERROR: 'adjudicante' column not found in XLSX")
        print(f"  Available columns: {list(header_map.keys())}")
        wb.close()
        return {}

    nif_contracts: dict[str, list[dict]] = defaultdict(list)
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        adj_text = str(row[adj_idx]) if row[adj_idx] else ""
        nif, name = extract_nif_from_adjudicante(adj_text)

        if not nif:
            continue

        contract = {"nif": nif, "entity_name": name}
        if valor_idx is not None and row[valor_idx]:
            try:
                contract["valor"] = float(row[valor_idx])
            except (ValueError, TypeError):
                contract["valor"] = 0
        else:
            contract["valor"] = 0
        if data_idx is not None and row[data_idx]:
            contract["data"] = str(row[data_idx])[:10]
        else:
            contract["data"] = ""
        if tipo_idx is not None and row[tipo_idx]:
            contract["tipo"] = str(row[tipo_idx])
        else:
            contract["tipo"] = ""
        if desc_idx is not None and row[desc_idx]:
            contract["objeto"] = str(row[desc_idx])[:200]
        else:
            contract["objeto"] = ""
        if id_idx is not None and row[id_idx]:
            contract["contract_id"] = int(row[id_idx])
        else:
            contract["contract_id"] = None
        if link_idx is not None and row[link_idx]:
            contract["link_pecas_proc"] = str(row[link_idx]).strip()
        else:
            contract["link_pecas_proc"] = ""

        nif_contracts[nif].append(contract)

    wb.close()
    result = dict(nif_contracts)

    # Cache for future runs
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    print(f"  Cached index to {CACHE_PATH.name}")
    print(f"  Parsed {total} rows → {len(result)} entities with contracts")
    return result


def cross_reference(
    bep_entities: list[dict],
    contract_index: dict[str, list[dict]],
    entity_filter: str = "",
    nif_filter: str = "",
) -> list[dict]:
    """Cross-reference BEP entities with BASE.gov.pt contracts."""
    results = []

    for entity in bep_entities:
        nif = entity["nif"]
        if not nif:
            continue

        if nif_filter and nif != nif_filter:
            continue

        if entity_filter:
            searchable = f"{entity['display_name']} {entity['entidade']} {entity['organismo']}".lower()
            if entity_filter.lower() not in searchable:
                continue

        contracts = contract_index.get(nif, [])
        total_value = sum(c.get("valor", 0) for c in contracts)

        results.append({
            "bep_entity": entity["display_name"],
            "bep_entidade": entity["entidade"],
            "bep_organismo": entity["organismo"],
            "bep_id": entity["id"],
            "nif": nif,
            "bep_listings": entity["listing_count"],
            "base_contracts": len(contracts),
            "base_total_value": total_value,
            "contracts": sorted(contracts, key=lambda c: c.get("data", ""), reverse=True)[:10],
        })

    results.sort(key=lambda r: r["base_contracts"], reverse=True)
    return results


def print_report(results: list[dict], verbose: bool = False):
    """Print the cross-reference report."""
    if not results:
        print("No cross-reference matches found.")
        print("Note: NIF enrichment may still be running (merge_nifs.py)")
        return

    total_contracts = sum(r["base_contracts"] for r in results)
    total_value = sum(r["base_total_value"] for r in results)
    matched = sum(1 for r in results if r["base_contracts"] > 0)

    print(f"\n{'='*80}")
    print(f"  BEP × BASE.gov.pt Cross-Reference Report")
    print(f"{'='*80}")
    print(f"  BEP entities with NIF:     {len(results)}")
    print(f"  Matched to contracts:      {matched}")
    print(f"  Total contracts found:     {total_contracts}")
    print(f"  Total contract value:      €{total_value:,.2f}")
    print(f"{'='*80}\n")

    for r in results:
        marker = "🔗" if r["base_contracts"] > 0 else "  "
        print(f"{marker} {r['bep_entity']}")
        print(f"   NIF: {r['nif']}  |  BEP listings: {r['bep_listings']}  |  "
              f"BASE contracts: {r['base_contracts']}  |  "
              f"Value: €{r['base_total_value']:,.2f}")

        if r["base_contracts"] > 0 and verbose:
            print(f"   Recent contracts:")
            for c in r["contracts"][:5]:
                valor = f"€{c.get('valor', 0):,.2f}" if c.get("valor") else "?"
                cid = c.get("contract_id")
                detail_url = f"{BASE_DETAIL_URL}{cid}" if cid else ""
                link_proc = c.get("link_pecas_proc", "")
                print(f"     [{c.get('data', '?')}] {valor}  {c.get('tipo', '')}  "
                      f"{c.get('objeto', '')[:60]}")
                if detail_url:
                    print(f"       📋 {detail_url}")
                if link_proc:
                    print(f"       📎 {link_proc[:100]}")
        print()


def print_stats(results: list[dict]):
    """Print summary statistics."""
    matched = [r for r in results if r["base_contracts"] > 0]
    unmatched = [r for r in results if r["base_contracts"] == 0]

    print(f"\n=== Cross-Reference Statistics ===")
    print(f"  Total BEP entities with NIF: {len(results)}")
    print(f"  Matched to BASE contracts:   {len(matched)}")
    print(f"  No contracts found:          {len(unmatched)}")

    if matched:
        top = sorted(matched, key=lambda r: r["base_contracts"], reverse=True)[:10]
        print(f"\n  Top entities by contract count:")
        for r in top:
            print(f"    {r['base_contracts']:4d} contracts  €{r['base_total_value']:>14,.2f}  "
                  f"{r['bep_entity'][:50]}")

        total_value = sum(r["base_total_value"] for r in matched)
        print(f"\n  Total contract value: €{total_value:,.2f}")


def main():
    parser = argparse.ArgumentParser(
        description="BEP × BASE.gov.pt Cross-Reference Report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--entity", default="", help="Filter by entity name")
    parser.add_argument("--nif", default="", help="Filter by NIF number")
    parser.add_argument("--export", help="Export to JSON file")
    parser.add_argument("--stats", action="store_true", help="Show summary stats only")
    parser.add_argument("-v", "--verbose", action="store_true", help="Show contract details")

    args = parser.parse_args()

    if not DB_PATH.exists():
        print(f"ERROR: BEP database not found at {DB_PATH}")
        print("Run `bep_scraper.py collect` first.")
        sys.exit(1)

    if not XLSX_PATH.exists():
        print(f"ERROR: Contracts data not found at {XLSX_PATH}")
        print("Run `merge_nifs.py` first to download the dataset.")
        sys.exit(1)

    print("Loading BEP entities with NIFs...")
    bep_entities = load_bep_entities_with_nif(DB_PATH)
    print(f"  Found {len(bep_entities)} entities with NIFs")

    print("Building contract index from BASE.gov.pt data...")
    contract_index = build_contract_index(XLSX_PATH)

    print("Cross-referencing...")
    results = cross_reference(
        bep_entities, contract_index,
        entity_filter=args.entity,
        nif_filter=args.nif,
    )

    if args.stats:
        print_stats(results)
    else:
        print_report(results, verbose=args.verbose)

    if args.export:
        output = {
            "summary": {
                "bep_entities_with_nif": len(results),
                "matched_to_contracts": sum(1 for r in results if r["base_contracts"] > 0),
                "total_contracts": sum(r["base_contracts"] for r in results),
                "total_value": sum(r["base_total_value"] for r in results),
            },
            "results": results,
        }
        with open(args.export, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        print(f"\nExported to {args.export}")


if __name__ == "__main__":
    main()
